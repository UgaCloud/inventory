# views/reports.py
from django.shortcuts import *
from django.http import JsonResponse, HttpResponse
from django.contrib.auth.decorators import login_required
from django.core.paginator import Paginator
from datetime import date, timedelta, time
import json
import csv
import random
import decimal
import xlsxwriter
import io
import calendar
from .reports import InventoryReports
from django.utils import timezone
from app.models.transactions import *
from app.models.products import *
from app.models.customers import *
from app.models.finance import *
from app.models.human_resource import *
from django.db.models.functions import TruncDate, Coalesce
from django.core.serializers.json import DjangoJSONEncoder
from decimal import Decimal
import pandas as pd
from reportlab.lib.styles import *
import csv
import json
from datetime import date
from django.http import HttpResponse
from django.db.models import *
from decimal import Decimal, DecimalTuple
from django.utils import timezone
from reportlab.lib import colors
from reportlab.lib.pagesizes import *
from reportlab.platypus import *

from io import BytesIO
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as ExcelImage
import matplotlib.pyplot as plt
import matplotlib
matplotlib.use('Agg') 
import traceback
import xlwt
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from io import BytesIO
from reportlab.pdfgen import canvas
from reportlab.lib.units import inch
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from django.db.models.functions import *
from django.core.paginator import *
from django.views.decorators.http import *


# For Excel export (if using pandas/openpyxl)
try:
    import pandas as pd
except ImportError:
    pd = None
    print("Warning: pandas not installed. Excel export may not work properly.")

try:
    import openpyxl
except ImportError:
    openpyxl = None
    print("Warning: openpyxl not installed. Excel export may not work properly.")

# For PDF export (reportlab)
try:
    from reportlab.pdfgen import canvas
    from reportlab.lib.pagesizes import letter
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.lib import colors
except ImportError:
    print("Warning: reportlab not installed. PDF export may not work properly.")
    # Define placeholders to avoid ImportError
    class canvas:
        pass
    letter = None
    class SimpleDocTemplate:
        pass
    class Table:
        pass
    class TableStyle:
        pass
    class Paragraph:
        pass
    class getSampleStyleSheet:
        @staticmethod
        def getSampleStyleSheet():
            return {}
    colors = None



REVENUE_EXPR = ExpressionWrapper(
    F('quantity') * F('sale_price'),
    output_field=DecimalField(max_digits=16, decimal_places=2)
)

COST_EXPR = ExpressionWrapper(
    F('quantity') * F('unit_cost'),
    output_field=DecimalField(max_digits=16, decimal_places=2)
)

# Custom JSON encoder class (add this after imports)
class CustomJSONEncoder(DjangoJSONEncoder):
    """Custom JSON encoder to handle dates and decimals"""
    def default(self, obj):
        if isinstance(obj, (date, timezone.date)):
            return obj.isoformat()
        if isinstance(obj, date.date):
            return obj.isoformat()
        if isinstance(obj, decimal.Decimal):
            return float(obj)
        if hasattr(obj, '__dict__'):
            return str(obj)
        return super().default(obj)



# ============================================================================
# REPORTS DASHBOARD & OVERVIEW VIEWS
# ============================================================================

@login_required
def reports_dashboard(request):
    """Reports dashboard view"""
    context = {}
    return render(request, 'reports/dashboard.html', context)

@login_required
def reports_details(request):
    """Reports dashboard view"""
    context = {}
    return render(request, 'reports/details.html', context)


# ============================================================================
# PURCHASE REPORTS VIEWS
# ============================================================================
# Supplier Purchase Summary, Purchase Trend Analysis, Purchase Order Status,
# Item-wise Purchase Analysis, Expiry Tracking Report
# ============================================================================

@login_required
def purchase_details(request, report_type='monthly', period=None):
    """
    Dynamic purchase report details view
    """
    today = timezone.now().date()
    
    # Get date range based on report_type
    if report_type == 'monthly':
        if not period:
            period = today.strftime('%Y-%m')
        year, month = map(int, period.split('-'))
        start_date = date(year, month, 1)
        if month == 12:
            end_date = date(year + 1, 1, 1) - timedelta(days=1)
        else:
            end_date = date(year, month + 1, 1) - timedelta(days=1)
        period_label = start_date.strftime('%B %Y')
        
    elif report_type == 'quarterly':
        if not period:
            quarter = (today.month - 1) // 3 + 1
            period = f"{today.year}-Q{quarter}"
        year, quarter = period.split('-Q')
        year = int(year)
        quarter = int(quarter)
        start_month = (quarter - 1) * 3 + 1
        start_date = date(year, start_month, 1)
        if start_month + 2 <= 12:
            end_date = date(year, start_month + 3, 1) - timedelta(days=1)
        else:
            end_date = date(year + 1, 1, 1).date() - timedelta(days=1)
        period_label = f"Q{quarter} {year}"
        
    elif report_type == 'yearly':
        if not period:
            period = str(today.year)
        year = int(period)
        start_date = date(year, 1, 1)
        end_date = date(year, 12, 31)
        period_label = str(year)
    else:
        start_date = today - timedelta(days=30)
        end_date = today
        period_label = f"{start_date.strftime('%b %d')} - {end_date.strftime('%b %d, %Y')}"
    
    # Get all purchase orders in date range
    purchase_orders = PurchaseOrder.objects.filter(
        purchase_date__range=[start_date, end_date]
    ).select_related('supplier', 'store')
    
    # Calculate totals
    total_purchases = purchase_orders.aggregate(
        total=Sum('total_cost')
    )['total'] or Decimal('0')
    
    total_orders = purchase_orders.count()
    avg_order_value = total_purchases / total_orders if total_orders > 0 else Decimal('0')
    
    # 1. Supplier Purchase Summary
    supplier_summary = PurchaseOrder.objects.filter(
        purchase_date__range=[start_date, end_date]
    ).values(
        'supplier__id',
        'supplier__name', 
        'supplier__supplier_code',
        'supplier__payment_terms'
    ).annotate(
        total_purchases=Sum('total_cost'),
        total_orders=Count('id'),
        avg_order_value=Avg('total_cost'),
        last_order_date=Max('purchase_date')
    ).order_by('-total_purchases')
    
    # Add performance rating with FIXED logic
    for supplier in supplier_summary:
        supplier['performance_rating'] = calculate_supplier_performance_fixed(
            supplier['supplier__id'],
            start_date,
            end_date
        )
    
    # 2. Purchase Trend Analysis - FIXED growth calculation
    trend_data = get_purchase_trend_data_corrected(report_type, start_date, end_date)
    
    # 3. Purchase Order Status - FIXED for consistency
    po_status_data = get_po_status_data_consistent(start_date, end_date)
    
    # 4. Item-wise Purchase Analysis
    item_analysis = get_item_analysis_data_accurate(start_date, end_date)
    
    # 5. Expiry Tracking - FIXED status logic
    expiry_data = get_expiry_data_with_correct_status(start_date, end_date)
    
    # Calculate monthly growth
    monthly_growth = Decimal('0')
    if len(trend_data) >= 2:
        current_total = trend_data[-1]['purchase_amount'] if trend_data else Decimal('0')
        previous_total = trend_data[-2]['purchase_amount'] if len(trend_data) >= 2 else Decimal('0')
        if previous_total > Decimal('0'):
            monthly_growth = ((current_total - previous_total) / previous_total) * 100
    
    # Calculate expiry statistics
    expiry_stats = calculate_expiry_stats_correct(expiry_data)
    
    # Prepare chart data
    trend_chart_data = {
        'labels': [t['month'] for t in trend_data[-6:]] if trend_data else [],
        'data': [float(t['purchase_amount']) for t in trend_data[-6:]] if trend_data else [],
    }
    
    item_chart_data = {
        'labels': [item['product__name'][:15] + '...' if len(item['product__name']) > 15 else item['product__name'] 
                  for item in item_analysis[:6]],
        'data': [float(item['total_cost']) for item in item_analysis[:6]],
    }
    
    context = {
        'report_type': report_type,
        'period': period,
        'period_label': period_label,
        'start_date': start_date,
        'end_date': end_date,
        'total_purchases': total_purchases,
        'total_orders': total_orders,
        'avg_order_value': avg_order_value,
        'monthly_growth': monthly_growth,
        'report_id': f"PUR-{period.replace('-', '').replace('Q', '') if period else timezone.now().strftime('%Y%m%d')}",
        'supplier_summary': supplier_summary,
        'trend_data': trend_data,
        'po_status_data': po_status_data,
        'item_analysis': item_analysis,
        'expiry_data': expiry_data,
        'expiry_stats': expiry_stats,
        'trend_chart_data': json.dumps(trend_chart_data),
        'item_chart_data': json.dumps(item_chart_data),
        'generated_by': request.user.get_full_name() or request.user.username,
        'generated_date': timezone.now(),
    }
    
    return render(request, 'reports/purchase_details.html', context)


def calculate_supplier_performance_fixed(supplier_id, start_date, end_date):
    """
    FIXED supplier performance rating - UGX 9M order should not be 'Poor'
    """
    orders = PurchaseOrder.objects.filter(
        supplier_id=supplier_id,
        purchase_date__range=[start_date, end_date]
    )
    
    if not orders.exists():
        return 'No Data'
    
    total_purchases = orders.aggregate(total=Sum('total_cost'))['total'] or Decimal('0')
    total_orders = orders.count()
    completed_orders = orders.filter(status__in=['completed', 'COMPLETED']).count()
    
    # Calculate completion rate
    completion_rate = (completed_orders / total_orders * 100) if total_orders > 0 else 0
    
    # Calculate average order value
    avg_order = total_purchases / total_orders if total_orders > 0 else Decimal('0')
    
    # NEW FIXED LOGIC: High value orders get better rating
    if total_orders == 0:
        return 'No Data'
    
    # Score based on multiple factors
    score = 0
    
    # High total purchases = high score
    if total_purchases > 8000000:  # > 8M UGX
        score += 4
    elif total_purchases > 3000000:  # > 3M UGX
        score += 3
    elif total_purchases > 1000000:  # > 1M UGX
        score += 2
    elif total_purchases > 500000:  # > 500K UGX
        score += 1
    
    # High average order value = high score
    if avg_order > 5000000:  # > 5M UGX average
        score += 3
    elif avg_order > 1000000:  # > 1M UGX average
        score += 2
    elif avg_order > 500000:  # > 500K UGX average
        score += 1
    
    # Good completion rate = high score
    if completion_rate >= 90:
        score += 3
    elif completion_rate >= 75:
        score += 2
    elif completion_rate >= 50:
        score += 1
    
    # Determine rating based on score
    if score >= 8:
        return 'Excellent'
    elif score >= 6:
        return 'Very Good'
    elif score >= 4:
        return 'Good'
    elif score >= 2:
        return 'Average'
    else:
        return 'Poor'


def get_purchase_trend_data_corrected(report_type, start_date, end_date):
    """
    FIXED purchase trend analysis with correct growth calculations
    """
    if report_type == 'monthly':
        periods = []
        today = timezone.now().date()
        
        # Get last 6 months including current
        for i in range(5, -1, -1):
            month_date = start_date - timedelta(days=30*i)
            month_start = date(month_date.year, month_date.month, 1)
            
            if month_start.month == 12:
                month_end = date(month_start.year + 1, 1, 1) - timedelta(days=1)
            else:
                month_end = date(month_start.year, month_start.month + 1, 1) - timedelta(days=1)
            
            if month_end > today:
                month_end = today
            
            month_orders = PurchaseOrder.objects.filter(
                purchase_date__range=[month_start, month_end]
            )
            
            month_total = month_orders.aggregate(
                total=Sum('total_cost')
            )['total'] or Decimal('0')
            
            month_count = month_orders.count()
            month_avg = month_total / month_count if month_count > 0 else Decimal('0')
            
            periods.append({
                'month': month_start.strftime('%B %Y'),
                'purchase_amount': month_total,
                'orders': month_count,
                'avg_order_value': month_avg,
                'growth_rate': Decimal('0'),  # Will calculate below
            })
        
        # Calculate growth rates CORRECTLY
        for i in range(len(periods)):
            if i > 0:
                prev_total = periods[i-1]['purchase_amount']
                current_total = periods[i]['purchase_amount']
                
                if prev_total > Decimal('0'):
                    growth = ((current_total - prev_total) / prev_total) * 100
                    periods[i]['growth_rate'] = growth
                else:
                    # If previous was 0 and current has value, that's 100% growth
                    if current_total > Decimal('0'):
                        periods[i]['growth_rate'] = Decimal('100')
                    else:
                        periods[i]['growth_rate'] = Decimal('0')
        
        return periods
    
    return []


def get_po_status_data_consistent(start_date, end_date):
    """
    FIXED purchase order status data - ensures consistency
    """
    # Get ALL orders in date range
    orders = PurchaseOrder.objects.filter(
        purchase_date__range=[start_date, end_date]
    ).select_related('supplier')
    
    # Count by status - FIXED: Handle case-insensitive
    status_counts = {
        'pending': orders.filter(
            Q(status='pending') | Q(status='PENDING')
        ).count(),
        'in_progress': orders.filter(
            Q(status='in_progress') | Q(status='IN_PROGRESS') | Q(status='in progress')
        ).count(),
        'completed': orders.filter(
            Q(status='completed') | Q(status='COMPLETED')
        ).count(),
        'cancelled': orders.filter(
            Q(status='cancelled') | Q(status='CANCELLED')
        ).count(),
    }
    
    # Detailed order list - MUST match counts
    order_details = []
    total_amount = Decimal('0')
    total_items = 0
    
    for order in orders.order_by('-purchase_date'):
        # FIXED: Calculate payment status
        payment_status = 'pending'
        if order.status.lower() == 'completed':
            payment_status = 'paid'
        elif order.expected_date and order.expected_date < timezone.now().date():
            payment_status = 'overdue'
        
        order_details.append({
            'po_number': f"PO-{order.id}",
            'supplier': order.supplier.name if order.supplier else 'Unknown',
            'order_date': order.purchase_date,
            'due_date': order.expected_date,
            'amount': order.total_cost or Decimal('0'),
            'items': order.items.count(),
            'status': order.status.lower(),  # Normalize to lowercase
            'payment_status': payment_status,
        })
        
        # Calculate totals
        total_amount += order.total_cost or Decimal('0')
        total_items += order.items.count()
    
    return {
        'status_counts': status_counts,
        'order_details': order_details,
        'total_amount': total_amount,
        'total_items': total_items,
        'total_orders': len(order_details),
    }


def get_item_analysis_data_accurate(start_date, end_date):
    """
    Get accurate item-wise purchase analysis
    """
    items = PurchaseOrderItem.objects.filter(
        order__purchase_date__range=[start_date, end_date]
    ).select_related('product', 'product__category').values(
        'product__id',
        'product__name',
        'product__sku',
        'product__category__name'
    ).annotate(
        quantity_purchased=Sum('quantity'),
        total_cost=Sum(F('quantity') * F('unit_cost')),
        avg_unit_cost=Avg('unit_cost'),
    ).order_by('-total_cost')[:15]
    
    # Add stock information
    for item in items:
        try:
            product = Product.objects.get(id=item['product__id'])
            
            # Get actual stock from inventory
            inventory = product.inventories.first()
            if inventory:
                current_stock = inventory.quantity_in_stock
                reorder_level = inventory.reorder_level
            else:
                current_stock = 0
                reorder_level = 10
            
            # Verify with batch stock
            batch_stock = InventoryBatch.objects.filter(
                product=product
            ).aggregate(
                total=Sum('remaining_quantity')
            )['total'] or 0
            
            # Use whichever is more accurate
            actual_stock = batch_stock if batch_stock > current_stock else current_stock
            
            item['current_stock'] = actual_stock
            item['reorder_level'] = reorder_level
            
            # Determine stock status
            if actual_stock == 0:
                item['stock_status'] = 'out_of_stock'
            elif actual_stock <= reorder_level:
                item['stock_status'] = 'low_stock'
            else:
                item['stock_status'] = 'in_stock'
                
        except Product.DoesNotExist:
            item['current_stock'] = 0
            item['reorder_level'] = 10
            item['stock_status'] = 'unknown'
    
    return items


def get_expiry_data_with_correct_status(start_date, end_date):
    """
    Get expiry tracking data with CORRECT status logic
    """
    today = timezone.now().date()
    expiry_data = []
    
    # Get all batches, not just recent ones
    batches = InventoryBatch.objects.filter(
        expiry_date__isnull=False,
        remaining_quantity__gt=0
    ).select_related('product').order_by('expiry_date')
    
    for batch in batches:
        if batch.expiry_date:
            days_remaining = (batch.expiry_date - today).days
            
            # CORRECT STATUS DETERMINATION
            if days_remaining < 0:
                status = 'expired'
                days_text = f"Expired {abs(days_remaining)} days ago"
            elif days_remaining == 0:
                status = 'expiring_today'  # CRITICAL, not safe!
                days_text = '0 days (Today)'
            elif days_remaining <= 7:
                status = 'critical'
                days_text = f'{days_remaining} days'
            elif days_remaining <= 30:
                status = 'expiring_soon'
                days_text = f'{days_remaining} days'
            elif days_remaining <= 60:
                status = 'monitor'
                days_text = f'{days_remaining} days'
            else:
                status = 'safe'
                days_text = f'{days_remaining} days'
        else:
            status = 'no_expiry'
            days_text = 'N/A'
        
        total_value = batch.remaining_quantity * batch.unit_cost
        
        expiry_data.append({
            'item_code': batch.product.sku if batch.product and batch.product.sku else 'N/A',
            'item_name': batch.product.name if batch.product else 'Unknown',
            'batch_number': f"BATCH-{batch.id}",
            'current_stock': batch.remaining_quantity,
            'purchase_date': batch.created_at.date() if batch.created_at else today,
            'expiry_date': batch.expiry_date,
            'days_remaining': days_text,
            'unit_cost': batch.unit_cost,
            'total_value': total_value,
            'status': status,
            'days_numeric': days_remaining if batch.expiry_date else 9999,
        })
    
    return expiry_data


def calculate_expiry_stats_correct(expiry_data):
    """
    Calculate correct expiry statistics
    """
    stats = {
        'expiring_soon_count': 0,
        'expired_count': 0,
        'good_stock_count': 0,
        'total_value_at_risk': Decimal('0'),
        'potential_loss': Decimal('0'),
    }
    
    for item in expiry_data:
        status = item['status']
        value = item['total_value']
        
        if status in ['expired', 'expiring_today', 'critical', 'expiring_soon']:
            stats['expiring_soon_count'] += 1
            
            if status == 'expired':
                stats['expired_count'] += 1
                risk_multiplier = Decimal('1.0')  # 100% at risk
            elif status == 'expiring_today':
                risk_multiplier = Decimal('0.9')  # 90% at risk
            elif status == 'critical':
                risk_multiplier = Decimal('0.7')  # 70% at risk
            elif status == 'expiring_soon':
                risk_multiplier = Decimal('0.5')  # 50% at risk
            else:
                risk_multiplier = Decimal('0.3')  # 30% at risk
            
            stats['total_value_at_risk'] += value
            stats['potential_loss'] += value * risk_multiplier
        elif status == 'safe' or status == 'monitor':
            stats['good_stock_count'] += 1
    
    return stats



def export_purchase_csv(request):
    """Export comprehensive purchase report as CSV"""
    # Get filter parameters from request
    report_type = request.GET.get('report_type', 'monthly')
    period = request.GET.get('period', '')
    store_id = request.GET.get('store')
    
    # Set up response
    response = HttpResponse(content_type='text/csv')
    filename = f"purchase_report_{period or timezone.now().strftime('%Y%m%d')}.csv"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    writer = csv.writer(response)
    
    # Get date range based on report_type
    today = timezone.now().date()
    if report_type == 'monthly':
        if period:
            year, month = map(int, period.split('-'))
        else:
            year, month = today.year, today.month
        start_date = date(year, month, 1).date()
        if month == 12:
            end_date = date(year + 1, 1, 1).date() - timedelta(days=1)
        else:
            end_date = date(year, month + 1, 1).date() - timedelta(days=1)
        period_label = start_date.strftime('%B %Y')
        
    elif report_type == 'quarterly':
        if period:
            year, quarter = period.split('-Q')
            year = int(year)
            quarter = int(quarter)
        else:
            year = today.year
            quarter = (today.month - 1) // 3 + 1
        start_month = (quarter - 1) * 3 + 1
        start_date = date(year, start_month, 1).date()
        if start_month + 2 <= 12:
            end_date = date(year, start_month + 3, 1).date() - timedelta(days=1)
        else:
            end_date = date(year + 1, 1, 1).date() - timedelta(days=1)
        period_label = f"Q{quarter} {year}"
        
    else:  # yearly or custom
        if period and len(period) == 4:
            year = int(period)
        else:
            year = today.year
        start_date = date(year, 1, 1).date()
        end_date = date(year, 12, 31).date()
        period_label = str(year)
    
    # Write header information
    writer.writerow(['PURCHASE REPORT EXPORT'])
    writer.writerow(['Generated:', timezone.now().strftime('%Y-%m-%d %H:%M:%S')])
    writer.writerow(['Period:', period_label])
    writer.writerow(['Date Range:', f"{start_date} to {end_date}"])
    writer.writerow(['Generated By:', request.user.get_full_name() or request.user.username])
    writer.writerow([])
    
    # Get purchase orders for the period
    purchase_orders = PurchaseOrder.objects.filter(
        purchase_date__range=[start_date, end_date]
    )
    
    # Calculate summary data
    total_purchases = purchase_orders.aggregate(total=Sum('total_cost'))['total'] or Decimal('0')
    total_orders = purchase_orders.count()
    avg_order_value = total_purchases / total_orders if total_orders > 0 else Decimal('0')
    
    # 1. SUMMARY SECTION
    writer.writerow(['SUMMARY SECTION'])
    writer.writerow([])
    
    writer.writerow(['Total Purchases:', f"UGX {total_purchases:,.0f}"])
    writer.writerow(['Total Orders:', f"{total_orders}"])
    writer.writerow(['Average Order Value:', f"UGX {avg_order_value:,.0f}"])
    writer.writerow(['Date Range:', f"{start_date} to {end_date}"])
    writer.writerow(['Days in Period:', f"{(end_date - start_date).days + 1}"])
    writer.writerow([])
    
    # 2. SUPPLIER ANALYSIS
    writer.writerow(['SUPPLIER PURCHASE ANALYSIS'])
    writer.writerow(['Supplier Name', 'Supplier Code', 'Total Purchases (UGX)', 'Orders', 'Avg Order (UGX)', 
                     'Last Order', 'Payment Terms', 'Performance Rating'])
    
    supplier_summary = PurchaseOrder.objects.filter(
        purchase_date__range=[start_date, end_date]
    ).values(
        'supplier__name', 
        'supplier__supplier_code',
        'supplier__payment_terms'
    ).annotate(
        total_purchases=Sum('total_cost'),
        orders=Count('id'),
        avg_order=Avg('total_cost'),
        last_order=Max('purchase_date')
    ).order_by('-total_purchases')
    
    for supplier in supplier_summary:
        # Calculate performance (simplified)
        performance = 'Good'
        if supplier['orders'] > 10 and supplier['avg_order'] > 100000:
            performance = 'Excellent'
        elif supplier['orders'] < 3:
            performance = 'New'
        
        writer.writerow([
            supplier['supplier__name'] or 'N/A',
            supplier['supplier__supplier_code'] or 'N/A',
            f"{supplier['total_purchases']:,.0f}" if supplier['total_purchases'] else '0',
            supplier['orders'],
            f"{supplier['avg_order']:,.0f}" if supplier['avg_order'] else '0',
            supplier['last_order'].strftime('%Y-%m-%d') if supplier['last_order'] else 'N/A',
            supplier['supplier__payment_terms'] or 'Net 30',
            performance
        ])
    
    writer.writerow([])
    
    # 3. PURCHASE TREND ANALYSIS
    writer.writerow(['PURCHASE TREND ANALYSIS'])
    writer.writerow(['Period', 'Purchase Amount (UGX)', 'Orders', 'Avg Order Value (UGX)', 'Growth Rate'])
    
    # Get trend data for last 6 periods
    trend_data = []
    if report_type == 'monthly':
        # Get last 6 months including current
        for i in range(6, 0, -1):
            month_date = end_date - timedelta(days=30*i)
            month_start = date(month_date.year, month_date.month, 1).date()
            if month_start.month == 12:
                month_end = date(month_start.year + 1, 1, 1).date() - timedelta(days=1)
            else:
                month_end = date(month_start.year, month_start.month + 1, 1).date() - timedelta(days=1)
            
            month_orders = PurchaseOrder.objects.filter(
                purchase_date__range=[month_start, month_end]
            )
            month_total = month_orders.aggregate(total=Sum('total_cost'))['total'] or Decimal('0')
            month_count = month_orders.count()
            month_avg = month_total / month_count if month_count > 0 else Decimal('0')
            
            # Calculate growth rate
            growth = '0%'
            if i < 6:
                prev_total = trend_data[-1][1]
                if prev_total > 0:
                    growth_rate = ((month_total - prev_total) / prev_total) * 100
                    growth = f"{growth_rate:+.1f}%"
            
            writer.writerow([
                month_start.strftime('%B %Y'),
                f"{month_total:,.0f}",
                month_count,
                f"{month_avg:,.0f}",
                growth
            ])
            trend_data.append((month_start.strftime('%B %Y'), month_total))
    
    writer.writerow([])
    
    # 4. PURCHASE ORDER STATUS
    writer.writerow(['PURCHASE ORDER STATUS'])
    writer.writerow(['PO Number', 'Supplier', 'Order Date', 'Due Date', 'Amount (UGX)', 
                     'Items', 'Status', 'Payment Status'])
    
    po_details = PurchaseOrder.objects.filter(
        purchase_date__range=[start_date, end_date]
    ).select_related('supplier').order_by('-purchase_date')[:50]
    
    for order in po_details:
        payment_status = 'Pending'
        if order.status == 'completed':
            payment_status = 'Paid'
        elif order.expected_date and order.expected_date < today:
            payment_status = 'Overdue'
        
        writer.writerow([
            f"PO-{order.id}",
            order.supplier.name if order.supplier else 'N/A',
            order.purchase_date.strftime('%Y-%m-%d') if order.purchase_date else '',
            order.expected_date.strftime('%Y-%m-%d') if order.expected_date else 'N/A',
            f"{order.total_cost:,.0f}",
            order.items.count(),
            order.status.capitalize(),
            payment_status
        ])
    
    writer.writerow([])
    
    # 5. ITEM-WISE PURCHASE ANALYSIS
    writer.writerow(['ITEM-WISE PURCHASE ANALYSIS'])
    writer.writerow(['Item Code', 'Item Name', 'Category', 'Quantity Purchased', 
                     'Unit Cost (UGX)', 'Total Cost (UGX)', 'Avg Unit Cost (UGX)'])
    
    item_analysis = PurchaseOrderItem.objects.filter(
        order__purchase_date__range=[start_date, end_date]
    ).values(
        'product__sku',
        'product__name',
        'product__category__name'
    ).annotate(
        quantity_purchased=Sum('quantity'),
        total_cost=Sum(F('quantity') * F('unit_cost')),
        avg_unit_cost=Avg('unit_cost')
    ).order_by('-total_cost')[:20]
    
    for item in item_analysis:
        writer.writerow([
            item['product__sku'] or 'N/A',
            item['product__name'],
            item['product__category__name'] or 'Uncategorized',
            item['quantity_purchased'],
            f"{item['avg_unit_cost']:,.0f}" if item['avg_unit_cost'] else '0',
            f"{item['total_cost']:,.0f}" if item['total_cost'] else '0',
            f"{item['avg_unit_cost']:,.0f}" if item['avg_unit_cost'] else '0'
        ])
    
    writer.writerow([])
    
    # 6. EXPIRY TRACKING
    writer.writerow(['EXPIRY TRACKING REPORT'])
    writer.writerow(['Item Code', 'Item Name', 'Batch Number', 'Quantity', 
                     'Purchase Date', 'Expiry Date', 'Days Remaining', 
                     'Unit Cost (UGX)', 'Total Value (UGX)', 'Status'])
    
    # Get batches that will expire within next 90 days or already expired
    ninety_days_from_now = today + timedelta(days=90)
    expiry_batches = InventoryBatch.objects.filter(
        expiry_date__lte=ninety_days_from_now,
        remaining_quantity__gt=0
    ).select_related('product').order_by('expiry_date')[:50]
    
    for batch in expiry_batches:
        days_remaining = (batch.expiry_date - today).days if batch.expiry_date else None
        
        if days_remaining is None:
            status = 'No Expiry'
        elif days_remaining < 0:
            status = 'Expired'
        elif days_remaining <= 30:
            status = 'Expiring Soon'
        elif days_remaining <= 60:
            status = 'Monitor'
        else:
            status = 'Safe'
        
        writer.writerow([
            batch.product.sku if batch.product.sku else 'N/A',
            batch.product.name,
            f"BATCH-{batch.id}",
            batch.remaining_quantity,
            batch.created_at.strftime('%Y-%m-%d') if batch.created_at else '',
            batch.expiry_date.strftime('%Y-%m-%d') if batch.expiry_date else 'N/A',
            f"{days_remaining}" if days_remaining is not None else 'N/A',
            f"{batch.unit_cost:,.0f}",
            f"{batch.remaining_quantity * batch.unit_cost:,.0f}",
            status
        ])
    
    writer.writerow([])
    
    # 7. STORE PERFORMANCE
    writer.writerow(['STORE PURCHASE PERFORMANCE'])
    writer.writerow(['Store Name', 'Total Purchases (UGX)', '% of Total', 
                     'Orders', 'Avg Order (UGX)', 'Performance'])
    
    store_performance = PurchaseOrder.objects.filter(
        purchase_date__range=[start_date, end_date]
    ).values(
        'store__name'
    ).annotate(
        total_purchases=Sum('total_cost'),
        orders=Count('id')
    ).order_by('-total_purchases')
    
    for store in store_performance:
        percent = (store['total_purchases'] / total_purchases * 100) if total_purchases > 0 else 0
        avg_order = store['total_purchases'] / store['orders'] if store['orders'] > 0 else Decimal('0')
        
        # Determine performance
        if percent > 40:
            performance = 'Top Performer'
        elif percent > 20:
            performance = 'Good'
        elif percent > 5:
            performance = 'Average'
        else:
            performance = 'Low Volume'
        
        writer.writerow([
            store['store__name'] or 'Unknown',
            f"{store['total_purchases']:,.0f}",
            f"{percent:.1f}%",
            store['orders'],
            f"{avg_order:,.0f}",
            performance
        ])
    
    writer.writerow([])
    
    # 8. CATEGORY ANALYSIS
    writer.writerow(['CATEGORY PURCHASE ANALYSIS'])
    writer.writerow(['Category', 'Total Purchases (UGX)', '% of Total', 
                     'Items Purchased', 'Avg Item Cost (UGX)', 'Trend'])
    
    category_analysis = PurchaseOrderItem.objects.filter(
        order__purchase_date__range=[start_date, end_date]
    ).values(
        'product__category__name'
    ).annotate(
        total_cost=Sum(F('quantity') * F('unit_cost')),
        total_items=Sum('quantity'),
        avg_item_cost=Avg('unit_cost')
    ).order_by('-total_cost')
    
    for category in category_analysis:
        percent = (category['total_cost'] / total_purchases * 100) if total_purchases > 0 else 0
        writer.writerow([
            category['product__category__name'] or 'Uncategorized',
            f"{category['total_cost']:,.0f}",
            f"{percent:.1f}%",
            category['total_items'],
            f"{category['avg_item_cost']:,.0f}",
            'Stable'  # Simplified trend
        ])
    
    return response


def export_purchase_pdf(request):
    """Export purchase report as PDF"""
    # Get filter parameters
    report_type = request.GET.get('report_type', 'monthly')
    period = request.GET.get('period', '')
    store_id = request.GET.get('store')
    
    # Set up response
    response = HttpResponse(content_type='application/pdf')
    filename = f"purchase_report_{period or timezone.now().strftime('%Y%m%d')}.pdf"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    # Get date range based on report_type
    today = timezone.now().date()
    if report_type == 'monthly':
        if period:
            year, month = map(int, period.split('-'))
        else:
            year, month = today.year, today.month
        start_date = date(year, month, 1).date()
        if month == 12:
            end_date = date(year + 1, 1, 1).date() - timedelta(days=1)
        else:
            end_date = date(year, month + 1, 1).date() - timedelta(days=1)
        period_label = start_date.strftime('%B %Y')
        
    elif report_type == 'quarterly':
        if period:
            year, quarter = period.split('-Q')
            year = int(year)
            quarter = int(quarter)
        else:
            year = today.year
            quarter = (today.month - 1) // 3 + 1
        start_month = (quarter - 1) * 3 + 1
        start_date = date(year, start_month, 1).date()
        if start_month + 2 <= 12:
            end_date = date(year, start_month + 3, 1).date() - timedelta(days=1)
        else:
            end_date = date(year + 1, 1, 1).date() - timedelta(days=1)
        period_label = f"Q{quarter} {year}"
        
    else:  # yearly or custom
        if period and len(period) == 4:
            year = int(period)
        else:
            year = today.year
        start_date = date(year, 1, 1).date()
        end_date = date(year, 12, 31).date()
        period_label = str(year)
    
    # Create PDF document
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter, 
                            rightMargin=72, leftMargin=72, 
                            topMargin=72, bottomMargin=72)
    elements = []
    styles = getSampleStyleSheet()
    
    # Custom styles
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Title'],
        fontSize=16,
        spaceAfter=30,
        alignment=1  # Center aligned
    )
    
    header_style = ParagraphStyle(
        'CustomHeader',
        parent=styles['Heading2'],
        fontSize=12,
        spaceAfter=12,
        spaceBefore=20,
        textColor=colors.HexColor('#2E6DA4')
    )
    
    # Title
    title = Paragraph(f"PURCHASE REPORT - {period_label}", title_style)
    elements.append(title)
    
    # Report info
    info_text = f"""
    <b>Date Range:</b> {start_date.strftime('%Y-%m-%d')} to {end_date.strftime('%Y-%m-%d')}<br/>
    <b>Generated By:</b> {request.user.get_full_name() or request.user.username}<br/>
    <b>Generated On:</b> {timezone.now().strftime('%Y-%m-%d %H:%M:%S')}<br/>
    <b>Report Type:</b> {report_type.capitalize()}
    """
    info_para = Paragraph(info_text, styles['Normal'])
    elements.append(info_para)
    elements.append(Spacer(1, 20))
    
    # Get summary data
    purchase_orders = PurchaseOrder.objects.filter(
        purchase_date__range=[start_date, end_date]
    )
    
    total_purchases = purchase_orders.aggregate(total=Sum('total_cost'))['total'] or Decimal('0')
    total_orders = purchase_orders.count()
    avg_order_value = total_purchases / total_orders if total_orders > 0 else Decimal('0')
    
    # SUMMARY SECTION
    elements.append(Paragraph("Executive Summary", header_style))
    
    summary_data = [
        ['Metric', 'Value'],
        ['Total Purchases', f"UGX {total_purchases:,.0f}"],
        ['Total Orders', str(total_orders)],
        ['Average Order Value', f"UGX {avg_order_value:,.0f}"],
        ['Date Range', f"{start_date.strftime('%Y-%m-%d')} to {end_date.strftime('%Y-%m-%d')}"],
        ['Days in Period', str((end_date - start_date).days + 1)]
    ]
    
    summary_table = Table(summary_data, colWidths=[2.5*inch, 3*inch])
    summary_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2E6DA4')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 11),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#F2F2F2')),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
    ]))
    
    elements.append(summary_table)
    elements.append(Spacer(1, 20))
    
    # SUPPLIER ANALYSIS (Top 10)
    elements.append(Paragraph("Top Suppliers by Purchase Volume", header_style))
    
    supplier_summary = PurchaseOrder.objects.filter(
        purchase_date__range=[start_date, end_date]
    ).values(
        'supplier__name'
    ).annotate(
        total_purchases=Sum('total_cost'),
        orders=Count('id')
    ).order_by('-total_purchases')[:10]
    
    supplier_data = [['Supplier', 'Total Purchases (UGX)', '% of Total', 'Orders', 'Avg Order']]
    
    for supplier in supplier_summary:
        percent = (supplier['total_purchases'] / total_purchases * 100) if total_purchases > 0 else 0
        avg_order = supplier['total_purchases'] / supplier['orders'] if supplier['orders'] > 0 else Decimal('0')
        
        supplier_data.append([
            supplier['supplier__name'][:20] + '...' if supplier['supplier__name'] and len(supplier['supplier__name']) > 20 else supplier['supplier__name'] or 'Unknown',
            f"{supplier['total_purchases']:,.0f}",
            f"{percent:.1f}%",
            str(supplier['orders']),
            f"{avg_order:,.0f}"
        ])
    
    supplier_table = Table(supplier_data, colWidths=[2*inch, 1.2*inch, 0.8*inch, 0.7*inch, 1.2*inch])
    supplier_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4A90E2')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
        ('ALIGN', (1, 1), (-1, -1), 'RIGHT'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 9),
        ('FONTSIZE', (0, 1), (-1, -1), 8),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
        ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#F8F9FA')),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
    ]))
    
    elements.append(supplier_table)
    elements.append(Spacer(1, 20))
    
    # PURCHASE ORDER STATUS
    elements.append(Paragraph("Purchase Order Status Summary", header_style))
    
    status_counts = PurchaseOrder.objects.filter(
        purchase_date__range=[start_date, end_date]
    ).values('status').annotate(
        count=Count('id'),
        total_amount=Sum('total_cost')
    ).order_by('-total_amount')
    
    status_data = [['Status', 'Count', '% of Total', 'Amount (UGX)', 'Avg Amount']]
    
    for status in status_counts:
        percent = (status['count'] / total_orders * 100) if total_orders > 0 else 0
        avg_amount = status['total_amount'] / status['count'] if status['count'] > 0 else Decimal('0')
        
        status_data.append([
            status['status'].capitalize(),
            str(status['count']),
            f"{percent:.1f}%",
            f"{status['total_amount']:,.0f}",
            f"{avg_amount:,.0f}"
        ])
    
    status_table = Table(status_data, colWidths=[1.5*inch, 0.8*inch, 0.8*inch, 1.2*inch, 1.2*inch])
    status_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#36B9CC')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
        ('ALIGN', (1, 1), (-1, -1), 'RIGHT'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 9),
        ('FONTSIZE', (0, 1), (-1, -1), 8),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
        ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#F8F9FA')),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
    ]))
    
    elements.append(status_table)
    elements.append(Spacer(1, 20))
    
    # ITEM ANALYSIS (Top 10)
    elements.append(Paragraph("Top Items by Purchase Value", header_style))
    
    item_analysis = PurchaseOrderItem.objects.filter(
        order__purchase_date__range=[start_date, end_date]
    ).values(
        'product__name'
    ).annotate(
        quantity=Sum('quantity'),
        total_cost=Sum(F('quantity') * F('unit_cost'))
    ).order_by('-total_cost')[:10]
    
    item_data = [['Item Name', 'Quantity', 'Total Cost (UGX)', '% of Total', 'Avg Unit Cost']]
    
    for item in item_analysis:
        percent = (item['total_cost'] / total_purchases * 100) if total_purchases > 0 else 0
        avg_unit = item['total_cost'] / item['quantity'] if item['quantity'] > 0 else Decimal('0')
        
        item_data.append([
            item['product__name'][:25] + '...' if item['product__name'] and len(item['product__name']) > 25 else item['product__name'] or 'Unknown',
            str(item['quantity']),
            f"{item['total_cost']:,.0f}",
            f"{percent:.1f}%",
            f"{avg_unit:,.0f}"
        ])
    
    item_table = Table(item_data, colWidths=[2.5*inch, 0.8*inch, 1.2*inch, 0.8*inch, 1.2*inch])
    item_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1CC88A')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
        ('ALIGN', (1, 1), (-1, -1), 'RIGHT'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 9),
        ('FONTSIZE', (0, 1), (-1, -1), 8),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
        ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#F8F9FA')),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
    ]))
    
    elements.append(item_table)
    elements.append(Spacer(1, 20))
    
    # EXPIRY TRACKING
    elements.append(Paragraph("Expiry Tracking (Next 60 Days)", header_style))
    
    sixty_days_from_now = today + timedelta(days=60)
    expiry_batches = InventoryBatch.objects.filter(
        expiry_date__range=[today, sixty_days_from_now],
        remaining_quantity__gt=0
    ).select_related('product').order_by('expiry_date')[:10]
    
    expiry_data = [['Item Name', 'Batch', 'Qty', 'Expiry Date', 'Days Left', 'Value (UGX)']]
    
    for batch in expiry_batches:
        days_left = (batch.expiry_date - today).days if batch.expiry_date else None
        
        expiry_data.append([
            batch.product.name[:20] + '...' if len(batch.product.name) > 20 else batch.product.name,
            f"#{batch.id}",
            str(batch.remaining_quantity),
            batch.expiry_date.strftime('%Y-%m-%d') if batch.expiry_date else 'N/A',
            str(days_left) if days_left is not None else 'N/A',
            f"{batch.remaining_quantity * batch.unit_cost:,.0f}"
        ])
    
    if len(expiry_batches) == 0:
        elements.append(Paragraph("No items expiring within the next 60 days.", styles['Normal']))
    else:
        expiry_table = Table(expiry_data, colWidths=[2*inch, 0.7*inch, 0.6*inch, 1*inch, 0.7*inch, 1*inch])
        expiry_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#F6C23E')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.black),
            ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
            ('ALIGN', (2, 1), (2, -1), 'RIGHT'),
            ('ALIGN', (5, 1), (5, -1), 'RIGHT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 9),
            ('FONTSIZE', (0, 1), (-1, -1), 8),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
            ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#FFF8E1')),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ]))
        
        elements.append(expiry_table)
    
    elements.append(Spacer(1, 20))
    
    # RECENT PURCHASE ORDERS
    elements.append(Paragraph("Recent Purchase Orders", header_style))
    
    recent_orders = PurchaseOrder.objects.filter(
        purchase_date__range=[start_date, end_date]
    ).select_related('supplier', 'store').order_by('-purchase_date')[:15]
    
    if recent_orders:
        order_data = [['PO#', 'Supplier', 'Date', 'Amount (UGX)', 'Status']]
        
        for order in recent_orders:
            order_data.append([
                f"PO-{order.id}",
                order.supplier.name[:15] + '...' if order.supplier and len(order.supplier.name) > 15 else order.supplier.name or 'N/A',
                order.purchase_date.strftime('%Y-%m-%d') if order.purchase_date else '',
                f"{order.total_cost:,.0f}",
                order.status.capitalize()
            ])
        
        order_table = Table(order_data, colWidths=[0.8*inch, 2*inch, 1*inch, 1.2*inch, 1*inch])
        order_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#6F42C1')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
            ('ALIGN', (3, 1), (3, -1), 'RIGHT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 8),
            ('FONTSIZE', (0, 1), (-1, -1), 7),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 6),
            ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#F8F9FA')),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ]))
        
        elements.append(order_table)
    else:
        elements.append(Paragraph("No purchase orders found in this period.", styles['Normal']))
    
    # Build PDF
    doc.build(elements)
    
    # Get PDF content and return response
    pdf = buffer.getvalue()
    buffer.close()
    response.write(pdf)
    
    return response


def export_purchase_excel(request):
    """Export comprehensive purchase report as Excel"""
    # Get filter parameters
    report_type = request.GET.get('report_type', 'monthly')
    period = request.GET.get('period', '')
    store_id = request.GET.get('store')
    
    # Get date range based on report_type
    today = timezone.now().date()
    if report_type == 'monthly':
        if period:
            year, month = map(int, period.split('-'))
        else:
            year, month = today.year, today.month
        start_date = date(year, month, 1).date()
        if month == 12:
            end_date = date(year + 1, 1, 1).date() - timedelta(days=1)
        else:
            end_date = date(year, month + 1, 1).date() - timedelta(days=1)
        period_label = start_date.strftime('%B %Y')
        
    elif report_type == 'quarterly':
        if period:
            year, quarter = period.split('-Q')
            year = int(year)
            quarter = int(quarter)
        else:
            year = today.year
            quarter = (today.month - 1) // 3 + 1
        start_month = (quarter - 1) * 3 + 1
        start_date = date(year, start_month, 1).date()
        if start_month + 2 <= 12:
            end_date = date(year, start_month + 3, 1).date() - timedelta(days=1)
        else:
            end_date = date(year + 1, 1, 1).date() - timedelta(days=1)
        period_label = f"Q{quarter} {year}"
        
    else:  # yearly or custom
        if period and len(period) == 4:
            year = int(period)
        else:
            year = today.year
        start_date = date(year, 1, 1).date()
        end_date = date(year, 12, 31).date()
        period_label = str(year)
    
    # Create Excel workbook
    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # Remove default sheet
    
    # Define styles
    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    subheader_font = Font(bold=True, color="000000", size=11)
    subheader_fill = PatternFill(start_color="C5D9F1", end_color="C5D9F1", fill_type="solid")
    
    total_font = Font(bold=True, color="000000", size=10)
    total_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    
    currency_format = '"UGX "#,##0'
    percent_format = '0.0"%'
    
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Get summary data
    purchase_orders = PurchaseOrder.objects.filter(
        purchase_date__range=[start_date, end_date]
    )
    
    total_purchases = purchase_orders.aggregate(total=Sum('total_cost'))['total'] or Decimal('0')
    total_orders = purchase_orders.count()
    avg_order_value = total_purchases / total_orders if total_orders > 0 else Decimal('0')
    
    # 1. SUMMARY SHEET
    ws_summary = wb.create_sheet(title="Summary")
    
    # Header
    ws_summary.merge_cells('A1:F1')
    ws_summary['A1'] = f"PURCHASE REPORT - {period_label}"
    ws_summary['A1'].font = header_font
    ws_summary['A1'].fill = header_fill
    ws_summary['A1'].alignment = header_alignment
    
    # Report info
    ws_summary['A3'] = "Generated:"
    ws_summary['B3'] = timezone.now().strftime('%Y-%m-%d %H:%M:%S')
    ws_summary['A4'] = "Period:"
    ws_summary['B4'] = period_label
    ws_summary['A5'] = "Generated By:"
    ws_summary['B5'] = request.user.get_full_name() or request.user.username
    ws_summary['A6'] = "Date Range:"
    ws_summary['B6'] = f"{start_date} to {end_date}"
    ws_summary['A7'] = "Report Type:"
    ws_summary['B7'] = report_type.capitalize()
    
    # Key Metrics
    ws_summary.merge_cells('A9:F9')
    ws_summary['A9'] = "KEY METRICS"
    ws_summary['A9'].font = subheader_font
    ws_summary['A9'].fill = subheader_fill
    ws_summary['A9'].alignment = Alignment(horizontal="center")
    
    summary_data = [
        ['Metric', 'Value'],
        ['Total Purchases', f"UGX {total_purchases:,.0f}"],
        ['Total Orders', total_orders],
        ['Average Order Value', f"UGX {avg_order_value:,.0f}"],
        ['Days in Period', (end_date - start_date).days + 1],
        ['Average Daily Purchases', f"UGX {total_purchases/((end_date - start_date).days + 1):,.0f}" if total_purchases > 0 else "UGX 0"]
    ]
    
    for i, row in enumerate(summary_data, start=10):
        for j, value in enumerate(row, start=1):
            cell = ws_summary.cell(row=i, column=j, value=value)
            cell.border = thin_border
            if i == 10:  # Header row
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
            elif j == 2 and 'UGX' in str(value):
                cell.number_format = '"UGX "#,##0'
    
    # 2. SUPPLIER ANALYSIS SHEET
    ws_suppliers = wb.create_sheet(title="Supplier Analysis")
    ws_suppliers['A1'] = "SUPPLIER PURCHASE ANALYSIS"
    ws_suppliers['A1'].font = header_font
    ws_suppliers['A1'].fill = header_fill
    ws_suppliers.merge_cells('A1:H1')
    ws_suppliers['A1'].alignment = header_alignment
    
    supplier_headers = ['Supplier Name', 'Supplier Code', 'Total Purchases (UGX)', 'Orders', 
                       'Avg Order (UGX)', 'Last Order Date', 'Payment Terms', 'Performance Rating']
    for col, header in enumerate(supplier_headers, start=1):
        cell = ws_suppliers.cell(row=3, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = subheader_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
    
    supplier_summary = PurchaseOrder.objects.filter(
        purchase_date__range=[start_date, end_date]
    ).values(
        'supplier__name', 
        'supplier__supplier_code',
        'supplier__payment_terms'
    ).annotate(
        total_purchases=Sum('total_cost'),
        orders=Count('id'),
        avg_order=Avg('total_cost'),
        last_order=Max('purchase_date')
    ).order_by('-total_purchases')
    
    row = 4
    for supplier in supplier_summary:
        # Calculate performance
        performance = 'Good'
        if supplier['orders'] > 10 and supplier['avg_order'] > 100000:
            performance = 'Excellent'
        elif supplier['orders'] >= 5 and supplier['avg_order'] > 50000:
            performance = 'Very Good'
        elif supplier['orders'] < 3:
            performance = 'New'
        elif supplier['orders'] > 0 and supplier['avg_order'] < 10000:
            performance = 'Needs Review'
        
        ws_suppliers.cell(row=row, column=1, value=supplier['supplier__name'] or 'N/A')
        ws_suppliers.cell(row=row, column=2, value=supplier['supplier__supplier_code'] or 'N/A')
        ws_suppliers.cell(row=row, column=3, value=float(supplier['total_purchases'] or 0))
        ws_suppliers.cell(row=row, column=4, value=supplier['orders'])
        ws_suppliers.cell(row=row, column=5, value=float(supplier['avg_order'] or 0))
        ws_suppliers.cell(row=row, column=6, value=supplier['last_order'].strftime('%Y-%m-%d') if supplier['last_order'] else 'N/A')
        ws_suppliers.cell(row=row, column=7, value=supplier['supplier__payment_terms'] or 'Net 30')
        ws_suppliers.cell(row=row, column=8, value=performance)
        
        # Format currency cells
        ws_suppliers.cell(row=row, column=3).number_format = currency_format
        ws_suppliers.cell(row=row, column=5).number_format = currency_format
        
        row += 1
    
    # Add totals row
    ws_suppliers.cell(row=row, column=1, value="TOTAL").font = total_font
    ws_suppliers.cell(row=row, column=2, value="")
    ws_suppliers.cell(row=row, column=3, value=float(total_purchases)).font = total_font
    ws_suppliers.cell(row=row, column=4, value=total_orders).font = total_font
    ws_suppliers.cell(row=row, column=5, value=float(avg_order_value)).font = total_font
    ws_suppliers.cell(row=row, column=3).number_format = currency_format
    ws_suppliers.cell(row=row, column=5).number_format = currency_format
    
    # Add border to all data cells
    for r in range(3, row + 1):
        for c in range(1, 9):
            ws_suppliers.cell(row=r, column=c).border = thin_border
    
    # 3. PURCHASE ORDER STATUS SHEET
    ws_status = wb.create_sheet(title="PO Status")
    ws_status['A1'] = "PURCHASE ORDER STATUS"
    ws_status['A1'].font = header_font
    ws_status['A1'].fill = header_fill
    ws_status.merge_cells('A1:I1')
    ws_status['A1'].alignment = header_alignment
    
    status_headers = ['PO Number', 'Supplier', 'Store', 'Order Date', 'Due Date', 
                     'Amount (UGX)', 'Items', 'Status', 'Payment Status']
    for col, header in enumerate(status_headers, start=1):
        cell = ws_status.cell(row=3, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = subheader_fill
        cell.border = thin_border
    
    po_details = PurchaseOrder.objects.filter(
        purchase_date__range=[start_date, end_date]
    ).select_related('supplier', 'store').order_by('-purchase_date')
    
    row = 4
    for order in po_details:
        payment_status = 'Pending'
        if order.status == 'completed':
            payment_status = 'Paid'
        elif order.expected_date and order.expected_date < today:
            payment_status = 'Overdue'
        
        ws_status.cell(row=row, column=1, value=f"PO-{order.id}")
        ws_status.cell(row=row, column=2, value=order.supplier.name if order.supplier else 'N/A')
        ws_status.cell(row=row, column=3, value=order.store.name if order.store else 'N/A')
        ws_status.cell(row=row, column=4, value=order.purchase_date.strftime('%Y-%m-%d') if order.purchase_date else '')
        ws_status.cell(row=row, column=5, value=order.expected_date.strftime('%Y-%m-%d') if order.expected_date else 'N/A')
        ws_status.cell(row=row, column=6, value=float(order.total_cost))
        ws_status.cell(row=row, column=7, value=order.items.count())
        ws_status.cell(row=row, column=8, value=order.status.capitalize())
        ws_status.cell(row=row, column=9, value=payment_status)
        
        # Format currency cell
        ws_status.cell(row=row, column=6).number_format = currency_format
        
        row += 1
    
    # Add border to all data cells
    for r in range(3, row):
        for c in range(1, 10):
            ws_status.cell(row=r, column=c).border = thin_border
    
    # 4. ITEM ANALYSIS SHEET
    ws_items = wb.create_sheet(title="Item Analysis")
    ws_items['A1'] = "ITEM-WISE PURCHASE ANALYSIS"
    ws_items['A1'].font = header_font
    ws_items['A1'].fill = header_fill
    ws_items.merge_cells('A1:G1')
    ws_items['A1'].alignment = header_alignment
    
    item_headers = ['Item Code', 'Item Name', 'Category', 'Quantity Purchased', 
                   'Avg Unit Cost (UGX)', 'Total Cost (UGX)', '% of Total']
    for col, header in enumerate(item_headers, start=1):
        cell = ws_items.cell(row=3, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = subheader_fill
        cell.border = thin_border
    
    item_analysis = PurchaseOrderItem.objects.filter(
        order__purchase_date__range=[start_date, end_date]
    ).values(
        'product__sku',
        'product__name',
        'product__category__name'
    ).annotate(
        quantity_purchased=Sum('quantity'),
        total_cost=Sum(F('quantity') * F('unit_cost')),
        avg_unit_cost=Avg('unit_cost')
    ).order_by('-total_cost')
    
    row = 4
    for item in item_analysis:
        percent = (item['total_cost'] / total_purchases * 100) if total_purchases > 0 else 0
        
        ws_items.cell(row=row, column=1, value=item['product__sku'] or 'N/A')
        ws_items.cell(row=row, column=2, value=item['product__name'])
        ws_items.cell(row=row, column=3, value=item['product__category__name'] or 'Uncategorized')
        ws_items.cell(row=row, column=4, value=item['quantity_purchased'])
        ws_items.cell(row=row, column=5, value=float(item['avg_unit_cost'] or 0))
        ws_items.cell(row=row, column=6, value=float(item['total_cost'] or 0))
        ws_items.cell(row=row, column=7, value=percent / 100)  # Excel expects decimal for percentage
        
        # Format cells
        ws_items.cell(row=row, column=5).number_format = currency_format
        ws_items.cell(row=row, column=6).number_format = currency_format
        ws_items.cell(row=row, column=7).number_format = percent_format
        
        row += 1
    
    # Add border to all data cells
    for r in range(3, row):
        for c in range(1, 8):
            ws_items.cell(row=r, column=c).border = thin_border
    
    # 5. EXPIRY TRACKING SHEET
    ws_expiry = wb.create_sheet(title="Expiry Tracking")
    ws_expiry['A1'] = "EXPIRY TRACKING REPORT"
    ws_expiry['A1'].font = header_font
    ws_expiry['A1'].fill = header_fill
    ws_expiry.merge_cells('A1:J1')
    ws_expiry['A1'].alignment = header_alignment
    
    expiry_headers = ['Item Code', 'Item Name', 'Batch Number', 'Current Stock', 
                     'Purchase Date', 'Expiry Date', 'Days Remaining', 
                     'Unit Cost (UGX)', 'Total Value (UGX)', 'Status']
    for col, header in enumerate(expiry_headers, start=1):
        cell = ws_expiry.cell(row=3, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = subheader_fill
        cell.border = thin_border
    
    # Get batches that will expire within next 120 days
    hundred_twenty_days_from_now = today + timedelta(days=120)
    expiry_batches = InventoryBatch.objects.filter(
        expiry_date__lte=hundred_twenty_days_from_now,
        remaining_quantity__gt=0
    ).select_related('product').order_by('expiry_date')
    
    row = 4
    total_expiry_value = Decimal('0')
    
    for batch in expiry_batches:
        days_remaining = (batch.expiry_date - today).days if batch.expiry_date else None
        
        if days_remaining is None:
            status = 'No Expiry'
        elif days_remaining < 0:
            status = 'Expired'
        elif days_remaining <= 30:
            status = 'Critical'
        elif days_remaining <= 60:
            status = 'Warning'
        elif days_remaining <= 90:
            status = 'Monitor'
        else:
            status = 'Safe'
        
        batch_value = batch.remaining_quantity * batch.unit_cost
        total_expiry_value += batch_value
        
        ws_expiry.cell(row=row, column=1, value=batch.product.sku if batch.product.sku else 'N/A')
        ws_expiry.cell(row=row, column=2, value=batch.product.name)
        ws_expiry.cell(row=row, column=3, value=f"BATCH-{batch.id}")
        ws_expiry.cell(row=row, column=4, value=batch.remaining_quantity)
        ws_expiry.cell(row=row, column=5, value=batch.created_at.strftime('%Y-%m-%d') if batch.created_at else '')
        ws_expiry.cell(row=row, column=6, value=batch.expiry_date.strftime('%Y-%m-%d') if batch.expiry_date else 'N/A')
        ws_expiry.cell(row=row, column=7, value=days_remaining if days_remaining is not None else 'N/A')
        ws_expiry.cell(row=row, column=8, value=float(batch.unit_cost))
        ws_expiry.cell(row=row, column=9, value=float(batch_value))
        ws_expiry.cell(row=row, column=10, value=status)
        
        # Format currency cells
        ws_expiry.cell(row=row, column=8).number_format = currency_format
        ws_expiry.cell(row=row, column=9).number_format = currency_format
        
        # Color code based on status
        if status == 'Expired':
            fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
        elif status == 'Critical':
            fill = PatternFill(start_color="FF6666", end_color="FF6666", fill_type="solid")
        elif status == 'Warning':
            fill = PatternFill(start_color="FFCC99", end_color="FFCC99", fill_type="solid")
        elif status == 'Monitor':
            fill = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")
        else:
            fill = PatternFill(start_color="CCFFCC", end_color="CCFFCC", fill_type="solid")
        
        for c in range(1, 11):
            ws_expiry.cell(row=row, column=c).fill = fill
        
        row += 1
    
    # Add summary row
    ws_expiry.cell(row=row, column=1, value="TOTAL AT RISK").font = total_font
    ws_expiry.cell(row=row, column=2, value="")
    ws_expiry.cell(row=row, column=3, value="")
    ws_expiry.cell(row=row, column=4, value="")
    ws_expiry.cell(row=row, column=5, value="")
    ws_expiry.cell(row=row, column=6, value="")
    ws_expiry.cell(row=row, column=7, value="")
    ws_expiry.cell(row=row, column=8, value="")
    ws_expiry.cell(row=row, column=9, value=float(total_expiry_value)).font = total_font
    ws_expiry.cell(row=row, column=10, value="")
    ws_expiry.cell(row=row, column=9).number_format = currency_format
    
    # Add border to all data cells
    for r in range(3, row + 1):
        for c in range(1, 11):
            ws_expiry.cell(row=r, column=c).border = thin_border
    
    # 6. STORE PERFORMANCE SHEET
    ws_stores = wb.create_sheet(title="Store Performance")
    ws_stores['A1'] = "STORE PURCHASE PERFORMANCE"
    ws_stores['A1'].font = header_font
    ws_stores['A1'].fill = header_fill
    ws_stores.merge_cells('A1:F1')
    ws_stores['A1'].alignment = header_alignment
    
    store_headers = ['Store Name', 'Total Purchases (UGX)', '% of Total', 
                    'Orders', 'Avg Order (UGX)', 'Performance Rating']
    for col, header in enumerate(store_headers, start=1):
        cell = ws_stores.cell(row=3, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = subheader_fill
        cell.border = thin_border
    
    store_performance = PurchaseOrder.objects.filter(
        purchase_date__range=[start_date, end_date]
    ).values(
        'store__name'
    ).annotate(
        total_purchases=Sum('total_cost'),
        orders=Count('id')
    ).order_by('-total_purchases')
    
    row = 4
    for store in store_performance:
        percent = (store['total_purchases'] / total_purchases * 100) if total_purchases > 0 else 0
        avg_order = store['total_purchases'] / store['orders'] if store['orders'] > 0 else Decimal('0')
        
        # Determine performance
        if percent > 40:
            rating = 'A+'
        elif percent > 30:
            rating = 'A'
        elif percent > 20:
            rating = 'B+'
        elif percent > 10:
            rating = 'B'
        elif percent > 5:
            rating = 'C'
        else:
            rating = 'D'
        
        ws_stores.cell(row=row, column=1, value=store['store__name'] or 'Unknown')
        ws_stores.cell(row=row, column=2, value=float(store['total_purchases']))
        ws_stores.cell(row=row, column=3, value=percent / 100)  # Excel expects decimal
        ws_stores.cell(row=row, column=4, value=store['orders'])
        ws_stores.cell(row=row, column=5, value=float(avg_order))
        ws_stores.cell(row=row, column=6, value=rating)
        
        # Format cells
        ws_stores.cell(row=row, column=2).number_format = currency_format
        ws_stores.cell(row=row, column=3).number_format = percent_format
        ws_stores.cell(row=row, column=5).number_format = currency_format
        
        row += 1
    
    # Add totals row
    ws_stores.cell(row=row, column=1, value="TOTAL").font = total_font
    ws_stores.cell(row=row, column=2, value=float(total_purchases)).font = total_font
    ws_stores.cell(row=row, column=3, value=1).font = total_font  # 100%
    ws_stores.cell(row=row, column=4, value=total_orders).font = total_font
    ws_stores.cell(row=row, column=5, value=float(avg_order_value)).font = total_font
    ws_stores.cell(row=row, column=6, value="")
    ws_stores.cell(row=row, column=2).number_format = currency_format
    ws_stores.cell(row=row, column=3).number_format = percent_format
    ws_stores.cell(row=row, column=5).number_format = currency_format
    
    # Add border to all data cells
    for r in range(3, row + 1):
        for c in range(1, 7):
            ws_stores.cell(row=r, column=c).border = thin_border
    
    # Auto-adjust column widths
    for ws in wb.worksheets:
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    # Save to response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f"purchase_report_{period or timezone.now().strftime('%Y%m%d')}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    wb.save(response)
    return response




# ============================================================================
# SALES REPORTS VIEWS
# ============================================================================
# Daily Sales Summary, Customer-wise Sales, Product Sales Performance,
# Payment Method Analysis, Transaction Audit Trail, Store Sales Performance
# ============================================================================


@login_required
def sales_details(request, period=None):
    """
    Sales report details view similar to inventory_details
    """
    # Default to current month if no period specified
    today = timezone.now()
    if not period:
        period = today.strftime("%B %Y")
        start_date = today.replace(day=1)
        end_date = today
    else:
        # Parse period (e.g., "January 2024")
        try:
            month_year = period.split()
            month = date.strptime(month_year[0], "%B").month
            year = int(month_year[1])
            start_date = timezone.make_aware(date(year, month, 1))
            if month == 12:
                end_date = timezone.make_aware(date(year, month, 31))
            else:
                end_date = timezone.make_aware(date(year, month+1, 1)) - timedelta(days=1)
        except:
            start_date = today.replace(day=1)
            end_date = today

    date_range = f"{start_date.strftime('%Y-%m-%d')} - {end_date.strftime('%Y-%m-%d')}"

 
    
    # Filter sales for the period - ALL sales
    sales_in_period = Sales.objects.filter(
        sale_date__range=[start_date.date(), end_date.date()]
    )
    
    # Basic metrics
    total_sales = sales_in_period.aggregate(total=Sum('total_amount'))['total'] or Decimal('0')
    total_transactions = sales_in_period.count()
    
    
    # Calculate average daily sales
    days_in_period = max((end_date.date() - start_date.date()).days + 1, 1)
    avg_daily_sales = total_sales / Decimal(days_in_period) if total_sales else Decimal('0')
    avg_transaction_value = total_sales / Decimal(total_transactions) if total_transactions > 0 else Decimal('0')
    
    # Daily sales data for chart
    daily_sales_data = []
    current_date = start_date.date()
    while current_date <= end_date.date():
        daily_sales = Sales.objects.filter(sale_date=current_date)
        daily_total = daily_sales.aggregate(total=Sum('total_amount'))['total'] or Decimal('0')
        daily_transactions = daily_sales.count()
        
        if daily_transactions > 0:
            daily_sales_data.append({
                'date': current_date.strftime('%Y-%m-%d'),
                'day': current_date.strftime('%A'),
                'total_sales': daily_total,
                'transactions': daily_transactions,
                'avg_transaction': daily_total / Decimal(daily_transactions) if daily_transactions > 0 else Decimal('0')
            })
        current_date += timedelta(days=1)

    # Customer-wise sales
    customer_sales = Sales.objects.filter(
        sale_date__range=[start_date.date(), end_date.date()]
    ).values(
        'customer__id', 'customer__name', 'customer__company'
    ).annotate(
        total_spent=Sum('total_amount'),
        transactions=Count('id')
    ).order_by('-total_spent')
    

    # Calculate average order value for each customer
    customer_sales_list = []
    for customer in customer_sales:
        avg_order = customer['total_spent'] / Decimal(customer['transactions']) if customer['transactions'] > 0 else Decimal('0')
        customer_sales_list.append({
            'customer__id': customer['customer__id'],
            'customer__name': customer['customer__name'],
            'customer__company': customer['customer__company'],
            'total_spent': customer['total_spent'],
            'transactions': customer['transactions'],
            'avg_order': avg_order
        })

    
    product_performance = SalesItem.objects.filter(
        order__sale_date__range=[start_date.date(), end_date.date()]
    ).values(
        'product__id', 'product__name', 'product__sku', 'product__category__name'
    ).annotate(
        units_sold=Sum('quantity'),
        revenue=Sum(F('quantity') * F('sale_price'), output_field=DecimalField()),
        avg_price=Avg('sale_price')
    ).order_by('-revenue')
    
    product_performance_list = list(product_performance)
   

    # Payment method analysis - Get ALL sales with payment methods
    payment_analysis = Sales.objects.filter(
        sale_date__range=[start_date.date(), end_date.date()],
        payment_method__isnull=False
    ).values(
        'payment_method__id', 'payment_method__name'
    ).annotate(
        transactions=Count('id'),
        total_amount=Sum('total_amount')
    ).order_by('-transactions')
    


    # Calculate average transaction for payment methods
    payment_analysis_list = []
    for payment in payment_analysis:
        avg_transaction = payment['total_amount'] / Decimal(payment['transactions']) if payment['transactions'] > 0 else Decimal('0')
        payment_analysis_list.append({
            'payment_method__id': payment['payment_method__id'],
            'payment_method__name': payment['payment_method__name'],
            'transactions': payment['transactions'],
            'total_amount': payment['total_amount'],
            'avg_transaction': avg_transaction
        })

    # Store performance
    store_performance = Sales.objects.filter(
        sale_date__range=[start_date.date(), end_date.date()]
    ).values(
        'store__id', 'store__name'
    ).annotate(
        total_sales=Sum('total_amount'),
        transactions=Count('id')
    ).order_by('-total_sales')
    
   

    # Calculate average transaction for stores
    store_performance_list = []
    for store in store_performance:
        avg_transaction = store['total_sales'] / Decimal(store['transactions']) if store['transactions'] > 0 else Decimal('0')
        store_performance_list.append({
            'store__id': store['store__id'],
            'store__name': store['store__name'],
            'total_sales': store['total_sales'],
            'transactions': store['transactions'],
            'avg_transaction': avg_transaction
        })

    # Recent transactions for audit trail
    recent_transactions = Sales.objects.filter(
        sale_date__range=[start_date.date(), end_date.date()]
    ).select_related('customer', 'store', 'payment_method').order_by('-sale_date', '-id')[:10]

    # Get top performing day (if any data exists)
    best_day = max(daily_sales_data, key=lambda x: x['total_sales']) if daily_sales_data else None

    # Prepare data for charts
    # Daily Sales Chart - show all days even with zero sales
    if daily_sales_data:
        daily_chart_data = {
            'labels': [d['date'] for d in daily_sales_data],
            'data': [float(d['total_sales']) for d in daily_sales_data]
        }
    else:
        daily_chart_data = {'labels': [], 'data': []}
    
    # Product Performance Chart
    if product_performance_list:
        product_chart_data = {
            'labels': [p['product__name'][:15] + '...' if len(p['product__name']) > 15 else p['product__name'] 
                       for p in product_performance_list],
            'data': [float(p['revenue']) for p in product_performance_list]
        }
    else:
        product_chart_data = {'labels': [], 'data': []}
    
    # Payment Method Pie Chart
    if payment_analysis_list:
        payment_chart_data = {
            'labels': [p['payment_method__name'] for p in payment_analysis_list],
            'data': [p['transactions'] for p in payment_analysis_list]
        }
    else:
        payment_chart_data = {'labels': [], 'data': []}
    
    # Store Performance Chart
    if store_performance_list:
        store_chart_data = {
            'labels': [s['store__name'] for s in store_performance_list],
            'data': [float(s['total_sales']) for s in store_performance_list]
        }
    else:
        store_chart_data = {'labels': [], 'data': []}

    # Counts for summary cards
    total_customers = sales_in_period.exclude(
        customer__isnull=True
    ).values('customer').distinct().count()
    
    
    new_customers = Customer.objects.filter(
        created_at__range=[start_date, end_date]
    ).count()
    
    # Repeat customers
    repeat_customers = Sales.objects.filter(
        sale_date__range=[start_date.date(), end_date.date()]
    ).exclude(customer__isnull=True).values('customer').distinct().count()
    
    # Average customer value
    avg_customer_value = total_sales / Decimal(max(len(customer_sales_list), 1)) if customer_sales_list else Decimal('0')
    
    # Store counts
    total_stores = StoreLocation.objects.filter(is_active=True).count()

    context = {
        'period': period,
        'date_range': date_range,
        'report_id': f"SAL-{start_date.strftime('%Y%m')}",
        'generated_by': request.user.get_full_name() or request.user.username,
        
        # Metrics
        'total_sales': total_sales,
        'total_transactions': total_transactions,
        'avg_daily_sales': avg_daily_sales,
        'avg_transaction_value': avg_transaction_value,
        'best_day': best_day,
        
        # Chart data (JSON serialized)
        'daily_chart_data': json.dumps(daily_chart_data),
        'product_chart_data': json.dumps(product_chart_data),
        'payment_chart_data': json.dumps(payment_chart_data),
        'store_chart_data': json.dumps(store_chart_data),
        
        # Tabular data
        'daily_sales_data': daily_sales_data,
        'customer_sales_data': customer_sales_list,
        'product_performance_data': product_performance_list,
        'payment_analysis_data': payment_analysis_list,
        'store_performance_data': store_performance_list,
        'recent_transactions': recent_transactions,
        
        # Counts for summary cards
        'total_customers': total_customers,
        'new_customers': new_customers,
        'repeat_customers': repeat_customers,
        'avg_customer_value': avg_customer_value,
        
        # Store counts
        'total_stores': total_stores,
        
        # Additional calculated values for template
        'days_count': len(daily_sales_data),
    }
    return render(request, 'reports/sales_details.html', context)


def export_sales_csv(request):
    """Export comprehensive sales report as CSV"""
    # Get filter parameters from request
    date_from = request.GET.get('date_from')
    date_to = request.GET.get('date_to')
    store_id = request.GET.get('store')
    period = request.GET.get('period', '')
    
    # Set up response
    response = HttpResponse(content_type='text/csv')
    filename = f"sales_report_{period or timezone.now().strftime('%Y%m%d')}.csv"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    writer = csv.writer(response)
    
    # Write header information
    writer.writerow(['SALES REPORT EXPORT'])
    writer.writerow(['Generated:', timezone.now().strftime('%Y-%m-%d %H:%M:%S')])
    writer.writerow(['Period:', period])
    writer.writerow(['Generated By:', request.user.get_full_name() or request.user.username])
    writer.writerow([])
    
    # Get the same data as the sales_details view
    today = timezone.now()
    if date_from and date_to:
        try:
            start_date = timezone.make_aware(date.strptime(date_from, '%Y-%m-%d'))
            end_date = timezone.make_aware(date.strptime(date_to, '%Y-%m-%d'))
        except:
            start_date = today.replace(day=1)
            end_date = today
    else:
        start_date = today.replace(day=1)
        end_date = today
    
    # Filter sales for the period
    sales_in_period = Sales.objects.filter(
        sale_date__range=[start_date.date(), end_date.date()]
    )
    
    # 1. SUMMARY SECTION
    writer.writerow(['SUMMARY SECTION'])
    writer.writerow([])
    
    total_sales = sales_in_period.aggregate(total=Sum('total_amount'))['total'] or Decimal('0')
    total_transactions = sales_in_period.count()
    days_in_period = max((end_date.date() - start_date.date()).days + 1, 1)
    avg_daily_sales = total_sales / Decimal(days_in_period) if total_sales else Decimal('0')
    
    writer.writerow(['Total Sales:', f"UGX {total_sales:,.0f}"])
    writer.writerow(['Total Transactions:', f"{total_transactions}"])
    writer.writerow(['Average Daily Sales:', f"UGX {avg_daily_sales:,.0f}"])
    writer.writerow(['Date Range:', f"{start_date.date()} to {end_date.date()}"])
    writer.writerow([])
    
    # 2. DAILY SALES SUMMARY
    writer.writerow(['DAILY SALES SUMMARY'])
    writer.writerow(['Date', 'Day', 'Total Sales (UGX)', 'Transactions', 'Avg Transaction (UGX)'])
    
    current_date = start_date.date()
    while current_date <= end_date.date():
        daily_sales = Sales.objects.filter(sale_date=current_date)
        daily_total = daily_sales.aggregate(total=Sum('total_amount'))['total'] or Decimal('0')
        daily_transactions = daily_sales.count()
        avg_transaction = daily_total / Decimal(daily_transactions) if daily_transactions > 0 else Decimal('0')
        
        writer.writerow([
            current_date.strftime('%Y-%m-%d'),
            current_date.strftime('%A'),
            f"{daily_total:,.0f}",
            daily_transactions,
            f"{avg_transaction:,.0f}"
        ])
        current_date += timedelta(days=1)
    
    writer.writerow([])
    
    # 3. CUSTOMER-WISE SALES
    writer.writerow(['CUSTOMER-WISE SALES ANALYSIS'])
    writer.writerow(['Customer Name', 'Customer Type', 'Total Spent (UGX)', 'Transactions', 'Avg Order (UGX)', 'Segment'])
    
    customer_sales = Sales.objects.filter(
        sale_date__range=[start_date.date(), end_date.date()]
    ).values(
        'customer__name', 'customer__company'
    ).annotate(
        total_spent=Sum('total_amount'),
        transactions=Count('id')
    ).order_by('-total_spent')
    
    for customer in customer_sales:
        avg_order = customer['total_spent'] / Decimal(customer['transactions']) if customer['transactions'] > 0 else Decimal('0')
        customer_type = 'Corporate' if customer['customer__company'] else 'Individual'
        
        # Determine segment
        if customer['total_spent'] > 1000000:
            segment = 'VIP'
        elif customer['transactions'] > 5:
            segment = 'Loyal'
        else:
            segment = 'Regular'
        
        writer.writerow([
            customer['customer__name'] or 'Walk-in Customer',
            customer_type,
            f"{customer['total_spent']:,.0f}",
            customer['transactions'],
            f"{avg_order:,.0f}",
            segment
        ])
    
    writer.writerow([])
    
    # 4. PRODUCT PERFORMANCE
    writer.writerow(['PRODUCT SALES PERFORMANCE'])
    writer.writerow(['Product Name', 'Product Code', 'Category', 'Units Sold', 'Revenue (UGX)', 'Avg Price (UGX)', '% of Total'])
    
    product_performance = SalesItem.objects.filter(
        order__sale_date__range=[start_date.date(), end_date.date()]
    ).values(
        'product__name', 'product__sku', 'product__category__name'
    ).annotate(
        units_sold=Sum('quantity'),
        revenue=Sum(F('quantity') * F('sale_price'), output_field=DecimalField()),
        avg_price=Avg('sale_price')
    ).order_by('-revenue')
    
    for product in product_performance:
        percent = (product['revenue'] / total_sales * 100) if total_sales > 0 else 0
        
        writer.writerow([
            product['product__name'],
            product['product__sku'] or 'N/A',
            product['product__category__name'] or 'Uncategorized',
            product['units_sold'],
            f"{product['revenue']:,.0f}",
            f"{product['avg_price']:,.0f}",
            f"{percent:.1f}%"
        ])
    
    writer.writerow([])
    
    # 5. PAYMENT METHOD ANALYSIS
    writer.writerow(['PAYMENT METHOD ANALYSIS'])
    writer.writerow(['Payment Method', 'Transactions', '% of Total', 'Total Amount (UGX)', 'Avg Transaction (UGX)'])
    
    payment_analysis = Sales.objects.filter(
        sale_date__range=[start_date.date(), end_date.date()],
        payment_method__isnull=False
    ).values(
        'payment_method__name'
    ).annotate(
        transactions=Count('id'),
        total_amount=Sum('total_amount')
    ).order_by('-transactions')
    
    for payment in payment_analysis:
        percent = (payment['transactions'] / total_transactions * 100) if total_transactions > 0 else 0
        avg_transaction = payment['total_amount'] / Decimal(payment['transactions']) if payment['transactions'] > 0 else Decimal('0')
        
        writer.writerow([
            payment['payment_method__name'] or 'Unknown',
            payment['transactions'],
            f"{percent:.1f}%",
            f"{payment['total_amount']:,.0f}",
            f"{avg_transaction:,.0f}"
        ])
    
    writer.writerow([])
    
    # 6. STORE PERFORMANCE
    writer.writerow(['STORE SALES PERFORMANCE'])
    writer.writerow(['Store', 'Sales (UGX)', '% of Total', 'Transactions', 'Avg Transaction (UGX)', 'Performance'])
    
    store_performance = Sales.objects.filter(
        sale_date__range=[start_date.date(), end_date.date()]
    ).values(
        'store__name'
    ).annotate(
        total_sales=Sum('total_amount'),
        transactions=Count('id')
    ).order_by('-total_sales')
    
    for store in store_performance:
        percent = (store['total_sales'] / total_sales * 100) if total_sales > 0 else 0
        avg_transaction = store['total_sales'] / Decimal(store['transactions']) if store['transactions'] > 0 else Decimal('0')
        
        # Determine performance rating
        if percent > 30:
            performance = 'Excellent'
        elif percent > 15:
            performance = 'Good'
        elif percent > 5:
            performance = 'Average'
        else:
            performance = 'Needs Attention'
        
        writer.writerow([
            store['store__name'],
            f"{store['total_sales']:,.0f}",
            f"{percent:.1f}%",
            store['transactions'],
            f"{avg_transaction:,.0f}",
            performance
        ])
    
    writer.writerow([])
    
    # 7. TRANSACTION AUDIT TRAIL
    writer.writerow(['TRANSACTION AUDIT TRAIL'])
    writer.writerow(['Transaction ID', 'Date', 'Customer', 'Store', 'Items', 'Total (UGX)', 'Payment Method', 'Status', 'Receipt No'])
    
    recent_transactions = Sales.objects.filter(
        sale_date__range=[start_date.date(), end_date.date()]
    ).select_related('customer', 'store', 'payment_method').order_by('-sale_date', '-id')[:100]
    
    for transaction in recent_transactions:
        writer.writerow([
            transaction.id,
            transaction.sale_date.strftime('%Y-%m-%d') if transaction.sale_date else '',
            transaction.customer.name if transaction.customer else 'Walk-in',
            transaction.store.name if transaction.store else '',
            transaction.number_of_items or 0,
            f"{transaction.total_amount:,.0f}",
            transaction.payment_method.name if transaction.payment_method else 'Unknown',
            transaction.status,
            transaction.receipt_no or 'N/A'
        ])
    
    return response

def export_sales_pdf(request):
    """Export sales report as PDF"""
    # Get filter parameters
    date_from = request.GET.get('date_from')
    date_to = request.GET.get('date_to')
    store_id = request.GET.get('store')
    period = request.GET.get('period', '')
    
    # Set up response
    response = HttpResponse(content_type='application/pdf')
    filename = f"sales_report_{period or timezone.now().strftime('%Y%m%d')}.pdf"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    # Get the same data as the sales_details view
    today = timezone.now()
    if date_from and date_to:
        try:
            start_date = timezone.make_aware(date.strptime(date_from, '%Y-%m-%d'))
            end_date = timezone.make_aware(date.strptime(date_to, '%Y-%m-%d'))
        except:
            start_date = today.replace(day=1)
            end_date = today
    else:
        start_date = today.replace(day=1)
        end_date = today
    
    # Create PDF
    doc = SimpleDocTemplate(response, pagesize=letter)
    elements = []
    styles = getSampleStyleSheet()
    
    # Title
    title = Paragraph(f"Sales Report - {period}", styles['Title'])
    elements.append(title)
    
    # Date range
    date_info = f"Period: {start_date.date()} to {end_date.date()}"
    date_para = Paragraph(date_info, styles['Normal'])
    elements.append(date_para)
    
    # Generated info
    generated_by = f"Generated by: {request.user.get_full_name() or request.user.username}"
    generated_date = f"Generated on: {timezone.now().strftime('%Y-%m-%d %H:%M:%S')}"
    elements.append(Paragraph(generated_by, styles['Normal']))
    elements.append(Paragraph(generated_date, styles['Normal']))
    elements.append(Spacer(1, 20))
    
    # Get data
    sales_in_period = Sales.objects.filter(
        sale_date__range=[start_date.date(), end_date.date()]
    )
    
    # SUMMARY SECTION
    elements.append(Paragraph("Summary", styles['Heading2']))
    
    total_sales = sales_in_period.aggregate(total=Sum('total_amount'))['total'] or Decimal('0')
    total_transactions = sales_in_period.count()
    days_in_period = max((end_date.date() - start_date.date()).days + 1, 1)
    avg_daily_sales = total_sales / Decimal(days_in_period) if total_sales else Decimal('0')
    
    summary_data = [
        ['Metric', 'Value'],
        ['Total Sales', f"UGX {total_sales:,.0f}"],
        ['Total Transactions', f"{total_transactions}"],
        ['Average Daily Sales', f"UGX {avg_daily_sales:,.0f}"],
        ['Date Range', f"{start_date.date()} to {end_date.date()}"]
    ]
    
    summary_table = Table(summary_data)
    summary_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 12),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
    ]))
    
    elements.append(summary_table)
    elements.append(Spacer(1, 20))
    
    # TRANSACTION AUDIT TRAIL
    elements.append(Paragraph("Transaction Audit Trail", styles['Heading2']))
    
    recent_transactions = Sales.objects.filter(
        sale_date__range=[start_date.date(), end_date.date()]
    ).select_related('customer', 'store', 'payment_method').order_by('-sale_date', '-id')[:50]
    
    audit_data = [['ID', 'Date', 'Customer', 'Store', 'Items', 'Total', 'Payment', 'Status']]
    
    for transaction in recent_transactions:
        audit_data.append([
            str(transaction.id),
            transaction.sale_date.strftime('%Y-%m-%d') if transaction.sale_date else '',
            transaction.customer.name[:15] + '...' if transaction.customer and len(transaction.customer.name) > 15 else (transaction.customer.name if transaction.customer else 'Walk-in'),
            transaction.store.name[:10] + '...' if transaction.store and len(transaction.store.name) > 10 else (transaction.store.name if transaction.store else ''),
            str(transaction.number_of_items or 0),
            f"UGX {transaction.total_amount:,.0f}",
            transaction.payment_method.name[:10] if transaction.payment_method else 'Unknown',
            transaction.status[:10]
        ])
    
    # Add totals
    audit_data.append(['', '', '', '', 'Total:', f"UGX {total_sales:,.0f}", '', f"{total_transactions} txns"])
    
    audit_table = Table(audit_data)
    audit_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 8),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 6),
        ('BACKGROUND', (0, 1), (-1, -2), colors.beige),
        ('FONTSIZE', (0, 1), (-1, -2), 7),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
        ('BACKGROUND', (0, -1), (-1, -1), colors.lightgrey),
        ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
    ]))
    
    elements.append(audit_table)
    elements.append(Spacer(1, 20))
    
    # STORE PERFORMANCE
    elements.append(Paragraph("Store Performance", styles['Heading2']))
    
    store_performance = Sales.objects.filter(
        sale_date__range=[start_date.date(), end_date.date()]
    ).values(
        'store__name'
    ).annotate(
        total_sales=Sum('total_amount'),
        transactions=Count('id')
    ).order_by('-total_sales')[:10]
    
    store_data = [['Store', 'Sales (UGX)', '% of Total', 'Transactions', 'Avg Transaction']]
    
    for store in store_performance:
        percent = (store['total_sales'] / total_sales * 100) if total_sales > 0 else 0
        avg_transaction = store['total_sales'] / Decimal(store['transactions']) if store['transactions'] > 0 else Decimal('0')
        
        store_data.append([
            store['store__name'][:15] if store['store__name'] else 'Unknown',
            f"UGX {store['total_sales']:,.0f}",
            f"{percent:.1f}%",
            str(store['transactions']),
            f"UGX {avg_transaction:,.0f}"
        ])
    
    store_table = Table(store_data)
    store_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
    ]))
    
    elements.append(store_table)
    
    # Build PDF
    doc.build(elements)
    return response

def export_sales_excel(request):
    """Export comprehensive sales report as Excel"""
    # Get filter parameters
    date_from = request.GET.get('date_from')
    date_to = request.GET.get('date_to')
    store_id = request.GET.get('store')
    period = request.GET.get('period', '')
    
    # Get the same data as the sales_details view
    today = timezone.now()
    if date_from and date_to:
        try:
            start_date = timezone.make_aware(date.strptime(date_from, '%Y-%m-%d'))
            end_date = timezone.make_aware(date.strptime(date_to, '%Y-%m-%d'))
        except:
            start_date = today.replace(day=1)
            end_date = today
    else:
        start_date = today.replace(day=1)
        end_date = today
    
    # Create Excel workbook
    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # Remove default sheet
    
    # Define styles
    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    subheader_font = Font(bold=True, color="000000", size=11)
    subheader_fill = PatternFill(start_color="C5D9F1", end_color="C5D9F1", fill_type="solid")
    
    total_font = Font(bold=True, color="000000", size=10)
    total_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # 1. SUMMARY SHEET
    ws_summary = wb.create_sheet(title="Summary")
    
    # Header
    ws_summary.merge_cells('A1:H1')
    ws_summary['A1'] = f"SALES REPORT - {period}"
    ws_summary['A1'].font = header_font
    ws_summary['A1'].fill = header_fill
    ws_summary['A1'].alignment = header_alignment
    
    ws_summary['A3'] = "Generated:"
    ws_summary['B3'] = timezone.now().strftime('%Y-%m-%d %H:%M:%S')
    ws_summary['A4'] = "Period:"
    ws_summary['B4'] = period
    ws_summary['A5'] = "Generated By:"
    ws_summary['B5'] = request.user.get_full_name() or request.user.username
    ws_summary['A6'] = "Date Range:"
    ws_summary['B6'] = f"{start_date.date()} to {end_date.date()}"
    
    # Get summary data
    sales_in_period = Sales.objects.filter(
        sale_date__range=[start_date.date(), end_date.date()]
    )
    
    total_sales = sales_in_period.aggregate(total=Sum('total_amount'))['total'] or Decimal('0')
    total_transactions = sales_in_period.count()
    days_in_period = max((end_date.date() - start_date.date()).days + 1, 1)
    avg_daily_sales = total_sales / Decimal(days_in_period) if total_sales else Decimal('0')
    
    # Summary table
    ws_summary['A8'] = "KEY METRICS"
    ws_summary['A8'].font = subheader_font
    ws_summary['A8'].fill = subheader_fill
    
    summary_data = [
        ['Metric', 'Value'],
        ['Total Sales', f"UGX {total_sales:,.0f}"],
        ['Total Transactions', total_transactions],
        ['Average Daily Sales', f"UGX {avg_daily_sales:,.0f}"],
        ['Days in Period', days_in_period],
        ['Average Transaction Value', f"UGX {total_sales/total_transactions:,.0f}" if total_transactions > 0 else "UGX 0"]
    ]
    
    for i, row in enumerate(summary_data, start=9):
        for j, value in enumerate(row, start=1):
            cell = ws_summary.cell(row=i, column=j, value=value)
            cell.border = thin_border
            if i == 9:  # Header row
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
    
    # 2. DAILY SALES SHEET
    ws_daily = wb.create_sheet(title="Daily Sales")
    ws_daily['A1'] = "DAILY SALES SUMMARY"
    ws_daily['A1'].font = header_font
    ws_daily['A1'].fill = header_fill
    ws_daily.merge_cells('A1:E1')
    ws_daily['A1'].alignment = header_alignment
    
    daily_headers = ['Date', 'Day', 'Total Sales (UGX)', 'Transactions', 'Avg Transaction (UGX)']
    for col, header in enumerate(daily_headers, start=1):
        cell = ws_daily.cell(row=3, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = subheader_fill
        cell.border = thin_border
    
    row = 4
    current_date = start_date.date()
    while current_date <= end_date.date():
        daily_sales = Sales.objects.filter(sale_date=current_date)
        daily_total = daily_sales.aggregate(total=Sum('total_amount'))['total'] or Decimal('0')
        daily_transactions = daily_sales.count()
        avg_transaction = daily_total / Decimal(daily_transactions) if daily_transactions > 0 else Decimal('0')
        
        ws_daily.cell(row=row, column=1, value=current_date.strftime('%Y-%m-%d'))
        ws_daily.cell(row=row, column=2, value=current_date.strftime('%A'))
        ws_daily.cell(row=row, column=3, value=float(daily_total))
        ws_daily.cell(row=row, column=4, value=daily_transactions)
        ws_daily.cell(row=row, column=5, value=float(avg_transaction))
        
        # Format currency cells
        ws_daily.cell(row=row, column=3).number_format = '"UGX "#,##0'
        ws_daily.cell(row=row, column=5).number_format = '"UGX "#,##0'
        
        row += 1
        current_date += timedelta(days=1)
    
    # Add border to all data cells
    for r in range(3, row):
        for c in range(1, 6):
            ws_daily.cell(row=r, column=c).border = thin_border
    
    # 3. CUSTOMER SALES SHEET
    ws_customers = wb.create_sheet(title="Customer Sales")
    ws_customers['A1'] = "CUSTOMER-WISE SALES"
    ws_customers['A1'].font = header_font
    ws_customers['A1'].fill = header_fill
    ws_customers.merge_cells('A1:F1')
    ws_customers['A1'].alignment = header_alignment
    
    customer_headers = ['Customer Name', 'Customer Type', 'Total Spent (UGX)', 'Transactions', 'Avg Order (UGX)', 'Segment']
    for col, header in enumerate(customer_headers, start=1):
        cell = ws_customers.cell(row=3, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = subheader_fill
        cell.border = thin_border
    
    customer_sales = Sales.objects.filter(
        sale_date__range=[start_date.date(), end_date.date()]
    ).values(
        'customer__name', 'customer__company'
    ).annotate(
        total_spent=Sum('total_amount'),
        transactions=Count('id')
    ).order_by('-total_spent')
    
    row = 4
    for customer in customer_sales:
        avg_order = customer['total_spent'] / Decimal(customer['transactions']) if customer['transactions'] > 0 else Decimal('0')
        customer_type = 'Corporate' if customer['customer__company'] else 'Individual'
        
        # Determine segment
        if customer['total_spent'] > 1000000:
            segment = 'VIP'
        elif customer['transactions'] > 5:
            segment = 'Loyal'
        else:
            segment = 'Regular'
        
        ws_customers.cell(row=row, column=1, value=customer['customer__name'] or 'Walk-in Customer')
        ws_customers.cell(row=row, column=2, value=customer_type)
        ws_customers.cell(row=row, column=3, value=float(customer['total_spent']))
        ws_customers.cell(row=row, column=4, value=customer['transactions'])
        ws_customers.cell(row=row, column=5, value=float(avg_order))
        ws_customers.cell(row=row, column=6, value=segment)
        
        # Format currency cells
        ws_customers.cell(row=row, column=3).number_format = '"UGX "#,##0'
        ws_customers.cell(row=row, column=5).number_format = '"UGX "#,##0'
        
        row += 1
    
    # Add border to all data cells
    for r in range(3, row):
        for c in range(1, 7):
            ws_customers.cell(row=r, column=c).border = thin_border
    
    # 4. PRODUCT PERFORMANCE SHEET
    ws_products = wb.create_sheet(title="Product Performance")
    ws_products['A1'] = "PRODUCT SALES PERFORMANCE"
    ws_products['A1'].font = header_font
    ws_products['A1'].fill = header_fill
    ws_products.merge_cells('A1:G1')
    ws_products['A1'].alignment = header_alignment
    
    product_headers = ['Product Name', 'Product Code', 'Category', 'Units Sold', 'Revenue (UGX)', 'Avg Price (UGX)', '% of Total']
    for col, header in enumerate(product_headers, start=1):
        cell = ws_products.cell(row=3, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = subheader_fill
        cell.border = thin_border
    
    product_performance = SalesItem.objects.filter(
        order__sale_date__range=[start_date.date(), end_date.date()]
    ).values(
        'product__name', 'product__sku', 'product__category__name'
    ).annotate(
        units_sold=Sum('quantity'),
        revenue=Sum(F('quantity') * F('sale_price'), output_field=DecimalField()),
        avg_price=Avg('sale_price')
    ).order_by('-revenue')
    
    row = 4
    for product in product_performance:
        percent = (product['revenue'] / total_sales * 100) if total_sales > 0 else 0
        
        ws_products.cell(row=row, column=1, value=product['product__name'])
        ws_products.cell(row=row, column=2, value=product['product__sku'] or 'N/A')
        ws_products.cell(row=row, column=3, value=product['product__category__name'] or 'Uncategorized')
        ws_products.cell(row=row, column=4, value=product['units_sold'])
        ws_products.cell(row=row, column=5, value=float(product['revenue']))
        ws_products.cell(row=row, column=6, value=float(product['avg_price']))
        ws_products.cell(row=row, column=7, value=percent)
        
        # Format cells
        ws_products.cell(row=row, column=5).number_format = '"UGX "#,##0'
        ws_products.cell(row=row, column=6).number_format = '"UGX "#,##0'
        ws_products.cell(row=row, column=7).number_format = '0.0"%'
        
        row += 1
    
    # Add border to all data cells
    for r in range(3, row):
        for c in range(1, 8):
            ws_products.cell(row=r, column=c).border = thin_border
    
    # 5. TRANSACTION AUDIT SHEET
    ws_transactions = wb.create_sheet(title="Transactions")
    ws_transactions['A1'] = "TRANSACTION AUDIT TRAIL"
    ws_transactions['A1'].font = header_font
    ws_transactions['A1'].fill = header_fill
    ws_transactions.merge_cells('A1:I1')
    ws_transactions['A1'].alignment = header_alignment
    
    transaction_headers = ['ID', 'Date', 'Customer', 'Store', 'Items', 'Total (UGX)', 'Payment Method', 'Status', 'Receipt No']
    for col, header in enumerate(transaction_headers, start=1):
        cell = ws_transactions.cell(row=3, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = subheader_fill
        cell.border = thin_border
    
    recent_transactions = Sales.objects.filter(
        sale_date__range=[start_date.date(), end_date.date()]
    ).select_related('customer', 'store', 'payment_method').order_by('-sale_date', '-id')
    
    row = 4
    for transaction in recent_transactions:
        ws_transactions.cell(row=row, column=1, value=transaction.id)
        ws_transactions.cell(row=row, column=2, value=transaction.sale_date.strftime('%Y-%m-%d') if transaction.sale_date else '')
        ws_transactions.cell(row=row, column=3, value=transaction.customer.name if transaction.customer else 'Walk-in')
        ws_transactions.cell(row=row, column=4, value=transaction.store.name if transaction.store else '')
        ws_transactions.cell(row=row, column=5, value=transaction.number_of_items or 0)
        ws_transactions.cell(row=row, column=6, value=float(transaction.total_amount))
        ws_transactions.cell(row=row, column=7, value=transaction.payment_method.name if transaction.payment_method else 'Unknown')
        ws_transactions.cell(row=row, column=8, value=transaction.status)
        ws_transactions.cell(row=row, column=9, value=transaction.receipt_no or 'N/A')
        
        # Format currency cell
        ws_transactions.cell(row=row, column=6).number_format = '"UGX "#,##0'
        
        row += 1
    
    # Add border to all data cells
    for r in range(3, row):
        for c in range(1, 10):
            ws_transactions.cell(row=r, column=c).border = thin_border
    
    # Auto-adjust column widths
    for ws in wb.worksheets:
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    # Save to response
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    filename = f"sales_report_{period or timezone.now().strftime('%Y%m%d')}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    wb.save(response)
    return response


# ============================================================================
# INVENTORY REPORTS VIEWS
# ============================================================================
# Stock Level Report, Batch Expiry Report, Expired Stock Report,
# Inventory Valuation, Stock Aging Report, Real-time Stock Availability,
# Store-wise Stock Distribution
# ============================================================================

@login_required
def inventory_details(request):
    """Main inventory report view with real-time inventory data"""
    
    # Get filter parameters from request
    period = request.GET.get('period', 'Current')
    store_id = request.GET.get('store')
    category_id = request.GET.get('category')
    stock_status = request.GET.get('stock_status', 'all')
    
    # Set date range for reports
    today = timezone.now().date()
    
    # Generate report ID
    report_id = f"INV-{today.strftime('%Y%m')}-{today.strftime('%H%M')}"
    
    # 1. Get REAL Inventory Summary Data
    from app.models.products import Product, Inventory, Category, StoreLocation
    from app.models.transactions import InventoryBatch, StockAdjustment, StockMovement
    from decimal import Decimal
    
    # Total products count
    total_products = Product.objects.filter(is_active=True).count()
    
    # Get all inventories
    inventories_qs = Inventory.objects.select_related('product', 'store')
    
    if store_id:
        inventories_qs = inventories_qs.filter(store_id=store_id)
    
    if category_id:
        inventories_qs = inventories_qs.filter(product__category_id=category_id)
    
    # 2. Calculate Stock Status Summary
    low_stock_items = inventories_qs.filter(
        quantity_in_stock__lte=F('reorder_level'),
        quantity_in_stock__gt=0
    ).count()
    
    out_of_stock_items = inventories_qs.filter(quantity_in_stock=0).count()
    in_stock_items = inventories_qs.filter(quantity_in_stock__gt=F('reorder_level')).count()
    overstock_items = inventories_qs.filter(quantity_in_stock__gt=F('reorder_level') * 2).count()
    
    # 3. Get REAL Stock Level Data
    stock_level_data = []
    inventories = inventories_qs.select_related('product', 'product__category', 'store')[:100]
    
    for inv in inventories:
        product = inv.product
        store = inv.store
        
        # Calculate days of stock
        avg_monthly_sales = 10
        days_of_stock = (inv.quantity_in_stock / avg_monthly_sales * 30) if avg_monthly_sales > 0 else 0
        
        # Get last received date
        last_batch = InventoryBatch.objects.filter(
            product=product,
            store=store
        ).order_by('-received_date').first()
        
        last_received = last_batch.received_date.date() if last_batch else None
        
        # Determine stock status
        if inv.quantity_in_stock == 0:
            stock_status_text = 'Out of Stock'
            stock_status_class = 'danger'
            table_class = 'table-danger'
        elif inv.quantity_in_stock <= inv.reorder_level:
            stock_status_text = 'Low Stock'
            stock_status_class = 'warning'
            table_class = 'table-warning'
        elif inv.quantity_in_stock > inv.reorder_level * 2:
            stock_status_text = 'Overstock'
            stock_status_class = 'info'
            table_class = 'table-info'
        else:
            stock_status_text = 'In Stock'
            stock_status_class = 'success'
            table_class = ''
        
        stock_level_data.append({
            'sku': product.sku,
            'product_name': product.name,
            'category': product.category.name if product.category else 'Uncategorized',
            'current_stock': inv.quantity_in_stock,
            'reorder_level': inv.reorder_level,
            'stock_status': stock_status_text,
            'stock_status_class': stock_status_class,
            'table_class': table_class,
            'days_of_stock': int(days_of_stock),
            'last_received': last_received,
            'store_name': store.name
        })
    
    # 4. Get REAL Batch Expiry Data
    batch_expiry_data = []
    expiry_batches = InventoryBatch.objects.select_related(
        'product', 'store'
    ).filter(
        remaining_quantity__gt=0,
        expiry_date__isnull=False
    ).order_by('expiry_date')[:50]
    
    # Calculate expiry counts
    critical_count = 0
    warning_count = 0
    monitor_count = 0
    safe_count = 0
    
    for batch in expiry_batches:
        days_to_expiry = (batch.expiry_date - today).days if batch.expiry_date else None
        
        # Count for summary cards
        if days_to_expiry:
            if days_to_expiry <= 30:
                critical_count += 1
                expiry_status = 'Critical'
                expiry_class = 'danger'
                table_class = 'table-danger'
            elif days_to_expiry <= 60:
                warning_count += 1
                expiry_status = 'Warning'
                expiry_class = 'warning'
                table_class = 'table-warning'
            elif days_to_expiry <= 90:
                monitor_count += 1
                expiry_status = 'Monitor'
                expiry_class = 'info'
                table_class = ''
            else:
                safe_count += 1
                expiry_status = 'Safe'
                expiry_class = 'success'
                table_class = ''
        else:
            expiry_status = 'No Expiry'
            expiry_class = 'secondary'
            table_class = ''
        
        value_at_risk = float(batch.remaining_quantity * batch.unit_cost)
        
        batch_expiry_data.append({
            'batch_no': f"BATCH-{batch.id}",
            'product_name': batch.product.name,
            'sku': batch.product.sku,
            'manufacture_date': batch.received_date.date() if batch.received_date else None,
            'expiry_date': batch.expiry_date,
            'days_to_expiry': days_to_expiry,
            'current_stock': batch.remaining_quantity,
            'value_at_risk': value_at_risk,
            'status': expiry_status,
            'status_class': expiry_class,
            'table_class': table_class,
            'store_name': batch.store.name
        })
    
    # 5. Get REAL Expired Stock Data
    expired_stock_data = []
    expired_batches = InventoryBatch.objects.select_related(
        'product', 'store'
    ).filter(
        remaining_quantity__gt=0,
        expiry_date__lt=today
    )[:30]
    
    total_expired_value = 0
    total_expired_units = 0
    
    for batch in expired_batches:
        days_expired = (today - batch.expiry_date).days
        total_value = float(batch.remaining_quantity * batch.unit_cost)
        total_expired_value += total_value
        total_expired_units += batch.remaining_quantity
        
        expired_stock_data.append({
            'sku': batch.product.sku,
            'product_name': batch.product.name,
            'category': batch.product.category.name if batch.product.category else 'Uncategorized',
            'batch_no': f"BATCH-{batch.id}",
            'expiry_date': batch.expiry_date,
            'days_expired': days_expired,
            'expired_units': batch.remaining_quantity,
            'unit_cost': float(batch.unit_cost),
            'total_value': total_value,
            'store_name': batch.store.name
        })
    
    # 6. Calculate Inventory Valuation by Category
    inventory_valuation_data = []
    categories = Category.objects.all()
    total_inventory_value_decimal = Decimal('0')  # Keep as Decimal for calculations
    total_inventory_value_float = 0.0  # For JSON serialization
    
    for category in categories:
        category_products = Product.objects.filter(category=category, is_active=True)
        category_skus = category_products.count()
        
        total_units = 0
        total_cost_decimal = Decimal('0')
        
        for product in category_products:
            # Get inventory for this product
            product_inventory = Inventory.objects.filter(product=product)
            product_stock = product_inventory.aggregate(total=Sum('quantity_in_stock'))['total'] or 0
            
            # Get average cost from batches
            avg_cost = InventoryBatch.objects.filter(product=product).aggregate(
                avg_cost=Avg('unit_cost')
            )['avg_cost'] or Decimal('0')
            
            total_units += product_stock
            total_cost_decimal += avg_cost * Decimal(str(product_stock))
        
        if total_units > 0 and total_cost_decimal > Decimal('0'):
            avg_unit_cost = total_cost_decimal / Decimal(str(total_units))
            total_inventory_value_decimal += total_cost_decimal
            
            # Convert to float for template and JSON
            total_cost_float = float(total_cost_decimal)
            avg_unit_cost_float = float(avg_unit_cost)
            total_market_value_float = float(avg_unit_cost * Decimal('1.5') * Decimal(str(total_units)))
            
            inventory_valuation_data.append({
                'category_name': category.name,
                'skus': category_skus,
                'total_units': total_units,
                'avg_unit_cost': avg_unit_cost_float,
                'total_cost': total_cost_float,
                'avg_selling_price': float(avg_unit_cost * Decimal('1.5')),
                'total_market_value': total_market_value_float,
                'gross_margin': 50
            })
    
    # Convert total inventory value to float
    total_inventory_value_float = float(total_inventory_value_decimal)
    
    # 7. Get Real-time Stock Availability
    real_time_stock_data = []
    for inv in inventories_qs.select_related('product', 'product__category', 'store')[:20]:
        product = inv.product
        
        # Get committed stock
        from app.models.transactions import StockTransferItem
        committed_stock = StockTransferItem.objects.filter(
            product=product,
            stock_transfer__from_store=inv.store,
            stock_transfer__status__in=['pending', 'in_transit']
        ).aggregate(committed=Sum('quantity'))['committed'] or 0
        
        # Get in-transit stock
        in_transit = StockTransferItem.objects.filter(
            product=product,
            stock_transfer__to_store=inv.store,
            stock_transfer__status__in=['pending', 'in_transit']
        ).aggregate(in_transit=Sum('quantity'))['in_transit'] or 0
        
        available_for_sale = max(0, inv.quantity_in_stock - committed_stock)
        
        # Determine stock status
        if inv.quantity_in_stock == 0:
            stock_status = 'Out of Stock'
            status_class = 'danger'
        elif inv.quantity_in_stock <= inv.reorder_level:
            stock_status = 'Low Stock'
            status_class = 'warning'
        elif inv.quantity_in_stock > inv.reorder_level * 2:
            stock_status = 'Overstock'
            status_class = 'info'
        else:
            stock_status = 'In Stock'
            status_class = 'success'
        
        real_time_stock_data.append({
            'sku': product.sku,
            'product_name': product.name,
            'category': product.category.name if product.category else 'Uncategorized',
            'available_stock': inv.quantity_in_stock,
            'reserved_stock': committed_stock,
            'available_for_sale': available_for_sale,
            'in_transit': in_transit,
            'stock_status': stock_status,
            'status_class': status_class,
            'location': f"{inv.store.name}"
        })
    
    # 8. Get Store-wise Stock Distribution
    store_distribution_data = []
    stores = StoreLocation.objects.filter(is_active=True)
    
    for store in stores:
        store_inventories = Inventory.objects.filter(store=store)
        
        total_skus = store_inventories.values('product').distinct().count()
        total_units = store_inventories.aggregate(total=Sum('quantity_in_stock'))['total'] or 0
        
        # Calculate inventory value
        total_value_decimal = Decimal('0')
        for inv in store_inventories.select_related('product')[:100]:
            avg_cost = InventoryBatch.objects.filter(
                product=inv.product,
                store=store
            ).aggregate(avg_cost=Avg('unit_cost'))['avg_cost'] or Decimal('0')
            
            total_value_decimal += avg_cost * Decimal(str(inv.quantity_in_stock))
        
        total_value_float = float(total_value_decimal)
        
        # Calculate stock health
        low_stock_count = store_inventories.filter(
            quantity_in_stock__lte=F('reorder_level'),
            quantity_in_stock__gt=0
        ).count()
        
        out_of_stock_count = store_inventories.filter(quantity_in_stock=0).count()
        
        if out_of_stock_count > 10:
            stock_health = 'Critical'
            health_class = 'danger'
        elif low_stock_count > 5 or out_of_stock_count > 0:
            stock_health = 'Needs Review'
            health_class = 'warning'
        else:
            stock_health = 'Healthy'
            health_class = 'success'
        
        store_distribution_data.append({
            'store_name': store.name,
            'total_skus': total_skus,
            'total_units': total_units,
            'inventory_value': total_value_float,
            'avg_stock_per_sku': total_units / total_skus if total_skus > 0 else 0,
            'stock_health': stock_health,
            'health_class': health_class
        })
    
    # 9. Calculate percentages and totals
    total_all_value_float = sum(item['inventory_value'] for item in store_distribution_data)
    
    # Calculate store totals
    store_total_skus = sum(item['total_skus'] for item in store_distribution_data)
    store_total_units = sum(item['total_units'] for item in store_distribution_data)
    
    # Calculate inventory valuation totals
    total_units_all = sum(item['total_units'] for item in inventory_valuation_data)
    total_market_value_all = sum(item['total_market_value'] for item in inventory_valuation_data)
    
    # Calculate percentages for store distribution
    for item in store_distribution_data:
        if total_all_value_float > 0:
            item['percentage_of_total'] = (item['inventory_value'] / total_all_value_float) * 100
        else:
            item['percentage_of_total'] = 0
    
    # Calculate percentages for inventory valuation
    for item in inventory_valuation_data:
        if total_inventory_value_float > 0:
            item['percentage_of_total'] = (item['total_cost'] / total_inventory_value_float) * 100
        else:
            item['percentage_of_total'] = 0
    
    # 10. Prepare context
    context = {
        'period': period,
        'report_id': report_id,
        'total_products': total_products,
        
        # Stock level summary
        'in_stock_items': in_stock_items,
        'low_stock_items': low_stock_items,
        'out_of_stock_items': out_of_stock_items,
        'overstock_items': overstock_items,
        
        # Stock level data
        'stock_level_data': stock_level_data,
        
        # Batch expiry data
        'batch_expiry_data': batch_expiry_data,
        'critical_count': critical_count,
        'warning_count': warning_count,
        'monitor_count': monitor_count,
        'safe_count': safe_count,
        
        # Expired stock data
        'expired_stock_data': expired_stock_data,
        'total_expired_value': total_expired_value,
        'total_expired_units': total_expired_units,
        'expired_items_count': len(expired_stock_data),
        
        # Inventory valuation
        'inventory_valuation_data': inventory_valuation_data,
        'total_inventory_value': total_inventory_value_float or total_all_value_float,
        'total_units_all': total_units_all,
        'total_market_value_all': total_market_value_all,
        
        # Real-time stock
        'real_time_stock_data': real_time_stock_data,
        
        # Store distribution
        'store_distribution_data': store_distribution_data,
        'total_stores': stores.count(),
        'store_total_skus': store_total_skus,
        'store_total_units': store_total_units,
        
        # Chart data
        'stock_status_data': json.dumps({
            'labels': ['In Stock', 'Low Stock', 'Out of Stock', 'Overstock'],
            'data': [in_stock_items, low_stock_items, out_of_stock_items, overstock_items],
            'colors': ['#1CC88A', '#F6C23E', '#E74A3B', '#4A90E2']
        }, cls=DjangoJSONEncoder),
        
        'inventory_value_data': json.dumps({
            'labels': [item['category_name'] for item in inventory_valuation_data[:6]],
            'data': [item['total_cost'] for item in inventory_valuation_data[:6]]
        }, cls=DjangoJSONEncoder),
        
        'store_distribution_chart_data': json.dumps({
            'labels': [item['store_name'] for item in store_distribution_data],
            'data': [item['percentage_of_total'] for item in store_distribution_data]
        }, cls=DjangoJSONEncoder),
        
        'stock_aging_data': json.dumps({
            'labels': ['<15 Days', '15-30 Days', '30-60 Days', '60-90 Days', '90-180 Days', '>180 Days'],
            'data': [in_stock_items, low_stock_items, 0, 0, 0, 0]
        }, cls=DjangoJSONEncoder),
        
        # Additional info
        'date_range': f"As of {today.strftime('%b %d, %Y')}",
        'generated_by': request.user.get_full_name() or request.user.username,
    }
    
    return render(request, 'reports/inventory_details.html', context)

# ============================================================================
# TRANSFER & MOVEMENT REPORTS VIEWS
# ============================================================================
# Inter-store Transfer Summary, Transfer Request Status, Transfer History Audit,
# Department-wise Transfers
# ============================================================================

@login_required
def transfer_details(request):
    # Date range filtering
    date_range = request.GET.get('date_range', '')
    start_date = None
    end_date = None
    
    if date_range:
        try:
            start_date_str, end_date_str = date_range.split(' - ')
            start_date = date.strptime(start_date_str, '%Y-%m-%d').date()
            end_date = date.strptime(end_date_str, '%Y-%m-%d').date()
        except:
            # Default to current month
            start_date = timezone.now().replace(day=1).date()
            end_date = timezone.now().date()
    else:
        # Default to current month
        start_date = timezone.now().replace(day=1).date()
        end_date = timezone.now().date()
    
    # Filter StockTransfers by date range
    transfers = StockTransfer.objects.filter(
        transfer_date__range=[start_date, end_date]
    ).select_related('from_store', 'to_store', 'created_by')
    
    # Filter TransferRequests by date range
    transfer_requests = TransferRequest.objects.filter(
        request_date__date__range=[start_date, end_date]
    ).select_related('from_store', 'to_store', 'requested_by', 'department')
    
    # Inter-store Transfer Summary - Using StockTransfer model
    inter_store_summary = transfers.values(
        'from_store__name', 'to_store__name'
    ).annotate(
        total_transfers=Count('id'),
        total_items=Sum('items__quantity')
    ).order_by('-total_transfers')
    
    # Calculate total values for each route
    for item in inter_store_summary:
        # Get actual transfer objects for this route
        route_transfers = transfers.filter(
            from_store__name=item['from_store__name'],
            to_store__name=item['to_store__name']
        )
        
        total_value = Decimal('0')
        for transfer in route_transfers:
            total_value += transfer.total_value or Decimal('0')
        
        item['total_value'] = total_value
    
    # Transfer Request Status counts - Using TransferRequest model ONLY
    request_status_counts = {
        'pending': transfer_requests.filter(status='pending').count(),
        'approved': transfer_requests.filter(status='approved').count(),
        'rejected': transfer_requests.filter(status='rejected').count(),
        'fulfilled': transfer_requests.filter(status='fulfilled').count(),
    }
    
    # Stock Transfer Status counts - Separate from requests
    transfer_status_counts = {
        'pending': transfers.filter(status='pending').count(),
        'in_transit': transfers.filter(status='in_transit').count(),
        'completed': transfers.filter(status='completed').count(),
        'cancelled': transfers.filter(status='cancelled').count(),
    }
    
    # Department-wise Transfers
    department_transfers = {}
    departments = Department.objects.all()
    
    for dept in departments:
        dept_requests = transfer_requests.filter(department=dept)
        dept_transfers = transfers.filter(transfer_request__department=dept)
        
        initiated_count = dept_requests.count()
        
        total_items_moved = 0
        total_value = Decimal('0')
        
        for transfer in dept_transfers:
            total_items_moved += transfer.total_quantity or 0
            total_value += transfer.total_value or Decimal('0')
        
        department_transfers[dept.name] = {
            'transfers_initiated': initiated_count,
            'total_items_moved': total_items_moved,
            'total_value': total_value,
        }
    
    # Calculate totals
    total_items_moved = transfers.aggregate(
        total=Sum('items__quantity')
    )['total'] or 0
    
    total_value = sum((t.total_value or Decimal('0')) for t in transfers)
    
    context = {
        'start_date': start_date,
        'end_date': end_date,
        'total_transfers': transfers.count(),
        'total_items_moved': total_items_moved,
        'total_value': total_value,
        'inter_store_summary': inter_store_summary,
        'request_status_counts': request_status_counts,
        'transfer_requests': transfer_requests,
        'department_transfers': department_transfers,
        'report_id': f"TRF-{start_date.strftime('%Y%m')}-{transfers.count():03d}",
        'generated_by': request.user.get_full_name() or request.user.username,
        'last_updated': timezone.now(),
    }
    
    return render(request, 'reports/transfer_details.html', context)


# ============================================================================
# STOCK ADJUSTMENT REPORTS VIEWS
# ============================================================================
# Adjustment History, Reason-wise Adjustments, User-wise Adjustments,
# Batch Adjustment Summary
# ============================================================================


@login_required
def stockadj_details(request):
    """Stock Adjustment Reports with filtering and pagination"""
    
    # Get filter parameters
    start_date_param = request.GET.get('start_date')
    end_date_param = request.GET.get('end_date')
    date_range_param = request.GET.get('date_range')
    store_id = request.GET.get('store_id')
    status = request.GET.get('status')
    adjustment_type = request.GET.get('type')
    reason = request.GET.get('reason')
    user_id = request.GET.get('user_id')
    page = request.GET.get('page', 1)
    
    # Default date range: last 30 days
    today = timezone.now().date()
    
    # Handle date parsing - check multiple possible sources
    start_date = None
    end_date = None
    
    # First try: Check if date_range parameter exists (from datepicker)
    if date_range_param:
        try:
            # Format: "YYYY-MM-DD - YYYY-MM-DD"
            dates = date_range_param.split(' - ')
            if len(dates) == 2:
                start_date = date.strptime(dates[0].strip(), '%Y-%m-%d').date()
                end_date = date.strptime(dates[1].strip(), '%Y-%m-%d').date()
        except (ValueError, AttributeError):
            pass
    
    # Second try: Check individual start_date and end_date parameters
    if not start_date or not end_date:
        if start_date_param and end_date_param:
            try:
                start_date = date.strptime(start_date_param, '%Y-%m-%d').date()
                end_date = date.strptime(end_date_param, '%Y-%m-%d').date()
            except (ValueError, AttributeError):
                pass
    
    # Third: Use default if still not set
    if not start_date or not end_date:
        end_date = today
        start_date = end_date - timedelta(days=30)
    
    # Ensure end_date is not before start_date
    if end_date < start_date:
        start_date, end_date = end_date, start_date
    
    # Base queryset with date filtering
    adjustments = StockAdjustment.objects.filter(
        created_at__date__gte=start_date,
        created_at__date__lte=end_date
    ).select_related(
        'store', 'product', 'unit', 'created_by'
    )
    
    # Apply other filters
    if store_id and store_id != 'all':
        adjustments = adjustments.filter(store_id=store_id)
    
    if status and status != 'all':
        adjustments = adjustments.filter(status=status)
    
    if user_id and user_id != 'all':
        adjustments = adjustments.filter(created_by_id=user_id)
    
    # Handle reason filter - check for 'None' string and empty values
    if reason and reason != 'all':
        # Convert JavaScript 'None' to Python None for empty check
        if reason.lower() == 'none':
            adjustments = adjustments.filter(Q(reason__isnull=True) | Q(reason__exact=''))
        else:
            adjustments = adjustments.filter(reason__icontains=reason)
    
    # Determine adjustment type (increase/decrease)
    if adjustment_type and adjustment_type != 'all':
        if adjustment_type == 'increase':
            adjustments = adjustments.filter(quantity_change__gt=0)
        elif adjustment_type == 'decrease':
            adjustments = adjustments.filter(quantity_change__lt=0)
    
    # Get statistics
    total_adjustments = adjustments.count()
    
    # Get increases and decreases using separate queries
    total_increases = adjustments.filter(quantity_change__gt=0).count()
    total_decreases = adjustments.filter(quantity_change__lt=0).count()
    
    # Get net quantity
    net_quantity = adjustments.aggregate(
        net_quantity=Sum('quantity_change')
    )['net_quantity'] or 0
    
    # Calculate value impact
    total_value_impact = Decimal('0')
    for adj in adjustments:
        if adj.unit_cost:
            total_value_impact += Decimal(str(adj.quantity_change)) * Decimal(str(adj.unit_cost))
    
    # Calculate approval rate
    approved_count = adjustments.filter(status='approved').count()
    approval_rate = (approved_count / total_adjustments * 100) if total_adjustments > 0 else 0
    
    # Group by reason for reason-wise analysis
    reason_stats = adjustments.exclude(reason__isnull=True).exclude(reason__exact='').values('reason').annotate(
        count=Count('id'),
        total_quantity=Sum('quantity_change'),
        avg_quantity=Avg('quantity_change')
    ).order_by('-count')
    
    # Group by user for user-wise analysis
    user_stats = adjustments.values(
        'created_by__id', 
        'created_by__first_name', 
        'created_by__last_name',
        'created_by__username'
    ).annotate(
        total_adjustments=Count('id'),
        total_increases=Count(
            Case(
                When(quantity_change__gt=0, then=1),
                output_field=IntegerField()
            )
        ),
        total_decreases=Count(
            Case(
                When(quantity_change__lt=0, then=1),
                output_field=IntegerField()
            )
        ),
        net_quantity=Sum('quantity_change')
    ).order_by('-total_adjustments')
    
    # Group by batch (reference) for batch summary
    batch_stats = adjustments.exclude(reference__isnull=True).exclude(reference__exact='').annotate(
        day=TruncDay('created_at')
    ).values('reference').annotate(
        batch_date=F('day'),
        total_items=Count('id'),
        total_quantity=Sum('quantity_change'),
        unique_products=Count('product', distinct=True),
        initiator=Concat(
            F('created_by__first_name'), 
            Value(' '), 
            F('created_by__last_name'),
            output_field=CharField()
        )
    ).order_by('-batch_date')
    
    # Prepare chart data for daily adjustments
    daily_adjustments = []
    
    # Get daily counts using a simpler approach
    daily_counts = {}
    for adj in adjustments:
        day = adj.created_at.date()
        if day not in daily_counts:
            daily_counts[day] = {'increases': 0, 'decreases': 0, 'count': 0}
        
        daily_counts[day]['count'] += 1
        if adj.quantity_change > 0:
            daily_counts[day]['increases'] += 1
        elif adj.quantity_change < 0:
            daily_counts[day]['decreases'] += 1
    
    # Sort by date and limit to last 30 days
    sorted_days = sorted(daily_counts.keys())
    for day in sorted_days[-30:]:
        daily_adjustments.append({
            'day': day.strftime('%Y-%m-%d'),
            'count': daily_counts[day]['count'],
            'increases': daily_counts[day]['increases'],
            'decreases': daily_counts[day]['decreases']
        })
    
    # Paginate adjustments
    paginator = Paginator(adjustments.order_by('-created_at'), 50)
    try:
        adjustments_page = paginator.page(page)
    except:
        adjustments_page = paginator.page(1)
    
    # Get filter options
    stores = StoreLocation.objects.filter(is_active=True)
    users = User.objects.filter(is_active=True)
    
    # Common reasons for filter dropdown
    common_reasons = adjustments.exclude(reason__isnull=True).exclude(reason__exact='').values_list(
        'reason', flat=True
    ).distinct()[:10]
    
    # Prepare context
    context = {
        'adjustments': adjustments_page,
        'total_adjustments': total_adjustments,
        'total_increases': total_increases,
        'total_decreases': total_decreases,
        'net_quantity': net_quantity,
        'net_value_impact': total_value_impact,
        'approval_rate': approval_rate,
        
        'reason_stats': list(reason_stats),
        'user_stats': list(user_stats),
        'batch_stats': list(batch_stats),
        
        'stores': stores,
        'users': users,
        'common_reasons': list(common_reasons),
        
        'start_date': start_date.strftime('%Y-%m-%d'),
        'end_date': end_date.strftime('%Y-%m-%d'),
        'selected_store': store_id,
        'selected_status': status,
        'selected_type': adjustment_type,
        'selected_reason': reason if reason != 'None' else '',
        'selected_user': user_id,
        
        'daily_adjustments_json': json.dumps(daily_adjustments),
        
        'report_period': start_date.strftime('%B %Y'),
        'report_id': f"ADJ-{start_date.strftime('%Y')}-{adjustments.count():04d}",
        'today': today,
    }
    
    return render(request, 'reports/stockadj_details.html', context)

@require_GET
@login_required
def adjustment_details_api(request, adjustment_id):
    """API endpoint for adjustment details"""
    try:
        adjustment = StockAdjustment.objects.select_related(
            'store', 'product', 'unit', 'created_by'
        ).get(id=adjustment_id)
        
        data = {
            'id': adjustment.id,
            'reference': adjustment.reference or f'ADJ-{adjustment.id:06d}',
            'created_at': adjustment.created_at.strftime('%Y-%m-%d %H:%M'),
            'store': adjustment.store.name,
            'product_name': adjustment.product.name,
            'sku': adjustment.product.sku,
            'quantity_change': adjustment.quantity_change,
            'unit_cost': str(adjustment.unit_cost or 0),
            'value_impact': str(abs(adjustment.quantity_change * (adjustment.unit_cost or 0))),
            'reason': adjustment.reason,
            'created_by': adjustment.created_by.get_full_name() or adjustment.created_by.username,
            'status': adjustment.get_status_display(),
            'status_color': {
                'approved': 'success',
                'pending': 'warning',
                'applied': 'info',
                'cancelled': 'danger'
            }.get(adjustment.status, 'secondary'),
            'notes': adjustment.note,
        }
        
        return JsonResponse(data)
    except StockAdjustment.DoesNotExist:
        return JsonResponse({'error': 'Adjustment not found'}, status=404)


@require_GET
@login_required
def batch_details_api(request, batch_reference):
    """API endpoint for batch details"""
    try:
        adjustments = StockAdjustment.objects.filter(
            reference=batch_reference
        ).select_related('store', 'product', 'unit', 'created_by')
        
        if not adjustments.exists():
            return JsonResponse({'error': 'Batch not found'}, status=404)
        
        first_adj = adjustments.first()
        total_items = adjustments.count()
        total_quantity = adjustments.aggregate(total=Sum('quantity_change'))['total'] or 0
        total_value = sum(adj.quantity_change * (adj.unit_cost or 0) for adj in adjustments)
        
        items = []
        for adj in adjustments:
            items.append({
                'product_name': adj.product.name,
                'sku': adj.product.sku,
                'quantity_change': adj.quantity_change,
                'unit_cost': str(adj.unit_cost or 0),
                'value_impact': str(abs(adj.quantity_change * (adj.unit_cost or 0))),
            })
        
        data = {
            'reference': batch_reference,
            'date': first_adj.created_at.strftime('%Y-%m-%d'),
            'total_items': total_items,
            'total_quantity': total_quantity,
            'total_value_impact': str(abs(total_value)),
            'store': first_adj.store.name,
            'initiated_by': first_adj.created_by.get_full_name() or first_adj.created_by.username,
            'status': first_adj.get_status_display(),
            'status_color': {
                'approved': 'success',
                'pending': 'warning',
                'applied': 'info',
                'cancelled': 'danger'
            }.get(first_adj.status, 'secondary'),
            'items': items,
        }
        
        return JsonResponse(data)
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)


# ============================================================================
# FINANCIAL REPORTS VIEWS
# ============================================================================
# Cost of Goods Sold (COGS), Profit Margin Analysis, Revenue vs Cost Report,
# Accounts Receivable
# ============================================================================

@login_required
def financial_details(request):
    end_date = timezone.now().date()
    start_date = end_date - timedelta(days=90)

    date_range = request.GET.get('date_range')
    if date_range:
        try:
            s, e = date_range.split(' - ')
            start_date = timezone.datetime.strptime(s, "%Y-%m-%d").date()
            end_date = timezone.datetime.strptime(e, "%Y-%m-%d").date()
        except ValueError:
            pass

    store_id = request.GET.get('store')
    sales_filter = {'order__store_id': store_id} if store_id else {}
    purchase_filter = {}
    if store_id:
        purchase_filter['order__store_id'] = store_id

    # ---------------------
    # EXPRESSIONS
    # ---------------------
    REVENUE_EXPR = ExpressionWrapper(
        F('sale_price') * F('quantity'),
        output_field=DecimalField(max_digits=12, decimal_places=2)
    )
    
    COST_EXPR = ExpressionWrapper(
        F('unit_cost') * F('quantity'),
        output_field=DecimalField(max_digits=12, decimal_places=2)
    )

    # ---------------------
    # SALES (Revenue)
    # ---------------------
    sales_items = SalesItem.objects.filter(
        order__sale_date__range=[start_date, end_date],
        **sales_filter
    ).select_related('product', 'product__category', 'order')

    total_revenue = sales_items.aggregate(
        total=Sum(REVENUE_EXPR)
    )['total'] or Decimal('0')

    # ---------------------
    # PROPER COGS CALCULATION
    # Using average cost method for simplicity
    # ---------------------
    total_cogs = Decimal('0')
    product_cogs_map = {}
    
    # First, get all unique products sold
    sold_products = sales_items.values('product_id').distinct()
    
    for sold_product in sold_products:
        product_id = sold_product['product_id']
        
        # Get total quantity sold for this product
        product_sales = sales_items.filter(product_id=product_id).aggregate(
            total_quantity=Sum('quantity'),
            total_revenue=Sum(REVENUE_EXPR)
        )
        
        total_quantity_sold = product_sales['total_quantity'] or 0
        
        if total_quantity_sold > 0:
            # Get average purchase cost for this product
            # Look for purchases before the sale date range (FIFO principle)
            avg_cost_result = PurchaseOrderItem.objects.filter(
                product_id=product_id,
                order__purchase_date__lte=end_date  # Purchases up to the report end date
            ).aggregate(
                avg_cost=Avg('unit_cost')
            )
            
            avg_cost = avg_cost_result['avg_cost'] or Decimal('0')
            
            # Calculate COGS for this product
            product_cogs = avg_cost * total_quantity_sold
            total_cogs += product_cogs
            product_cogs_map[product_id] = {
                'avg_cost': float(avg_cost),
                'quantity_sold': total_quantity_sold,
                'cogs': float(product_cogs)
            }

    # If no proper COGS data, use estimated 60% COGS ratio
    if total_cogs == 0 and total_revenue > 0:
        total_cogs = total_revenue * Decimal('0.60')  # 60% COGS, 40% margin
    
    # Calculate financial metrics
    cogs_ratio = (total_cogs / total_revenue * 100) if total_revenue else Decimal('0')
    gross_profit = total_revenue - total_cogs
    gross_margin = (gross_profit / total_revenue * 100) if total_revenue else Decimal('0')

    # ---------------------
    # MONTHLY TRENDS with PROPER COGS
    # ---------------------
    monthly_trends = []
    
    # Get all months in the date range
    current_month = start_date.replace(day=1)
    while current_month <= end_date.replace(day=1):
        month_start = current_month
        month_end = (current_month + timedelta(days=32)).replace(day=1) - timedelta(days=1)
        
        # Revenue for this month
        month_sales = sales_items.filter(
            order__sale_date__range=[month_start, month_end]
        )
        
        month_revenue = month_sales.aggregate(
            total=Sum(REVENUE_EXPR)
        )['total'] or Decimal('0')
        
        # Calculate COGS for this month using product-level data
        month_cogs = Decimal('0')
        
        if month_revenue > 0:
            # Get unique products sold this month
            month_products = month_sales.values('product_id').distinct()
            
            for product in month_products:
                product_id = product['product_id']
                
                # Get quantity sold this month for this product
                product_month_sales = month_sales.filter(
                    product_id=product_id
                ).aggregate(
                    quantity=Sum('quantity')
                )
                
                quantity_sold = product_month_sales['quantity'] or 0
                
                if quantity_sold > 0:
                    # Use average cost from our product map
                    if product_id in product_cogs_map:
                        avg_cost = Decimal(str(product_cogs_map[product_id]['avg_cost']))
                    else:
                        # Fallback to average purchase cost
                        avg_cost_result = PurchaseOrderItem.objects.filter(
                            product_id=product_id
                        ).aggregate(
                            avg_cost=Avg('unit_cost')
                        )
                        avg_cost = avg_cost_result['avg_cost'] or Decimal('0')
                    
                    month_cogs += avg_cost * quantity_sold
        
        # If no proper COGS, use estimated
        if month_cogs == 0 and month_revenue > 0:
            month_cogs = month_revenue * Decimal('0.60')
        
        # Calculate profit and margin
        month_profit = month_revenue - month_cogs
        month_margin = (month_profit / month_revenue * 100) if month_revenue else Decimal('0')
        
        monthly_trends.append({
            'month': current_month.strftime('%Y-%m'),
            'date': current_month.strftime('%b %Y'),
            'revenue': float(month_revenue),
            'cogs': float(month_cogs),
            'profit': float(month_profit),
            'margin': float(month_margin)
        })
        
        # Move to next month
        current_month = (current_month + timedelta(days=32)).replace(day=1)

    # Convert to dictionary for template
    monthly_trends_dict = {item['month']: item for item in monthly_trends}

    # ---------------------
    # CATEGORY PERFORMANCE with PROPER COGS
    # ---------------------
    category_performance = []
    
    # Get all categories from sales
    sales_by_category = sales_items.values(
        'product__category__id', 'product__category__name'
    ).annotate(
        revenue=Sum(REVENUE_EXPR),
        quantity=Sum('quantity')
    ).order_by('-revenue')
    
    for cat_data in sales_by_category:
        category_id = cat_data['product__category__id']
        category_name = cat_data['product__category__name'] or 'Uncategorized'
        revenue = cat_data['revenue'] or Decimal('0')
        quantity = cat_data['quantity'] or 0
        
        # Calculate COGS for this category using products in the category
        category_cogs = Decimal('0')
        
        if revenue > 0:
            # Get products in this category that were sold
            category_products = sales_items.filter(
                product__category_id=category_id
            ).values('product_id').distinct()
            
            for product in category_products:
                product_id = product['product_id']
                
                # Get quantity sold for this product in this category
                product_sales = sales_items.filter(
                    product_id=product_id,
                    product__category_id=category_id
                ).aggregate(
                    quantity=Sum('quantity')
                )
                
                product_quantity = product_sales['quantity'] or 0
                
                if product_quantity > 0:
                    # Use average cost from our product map
                    if product_id in product_cogs_map:
                        avg_cost = Decimal(str(product_cogs_map[product_id]['avg_cost']))
                    else:
                        # Fallback to average purchase cost
                        avg_cost_result = PurchaseOrderItem.objects.filter(
                            product_id=product_id
                        ).aggregate(
                            avg_cost=Avg('unit_cost')
                        )
                        avg_cost = avg_cost_result['avg_cost'] or Decimal('0')
                    
                    category_cogs += avg_cost * product_quantity
        
        # If no proper COGS, use estimated
        if category_cogs == 0 and revenue > 0:
            category_cogs = revenue * Decimal('0.60')
        
        profit = revenue - category_cogs
        margin = (profit / revenue * 100) if revenue else Decimal('0')
        
        category_performance.append({
            'category': {
                'id': category_id,
                'name': category_name
            },
            'revenue': float(revenue),
            'cogs': float(category_cogs),
            'profit': float(profit),
            'margin': float(margin),
            'quantity': quantity
        })

    # ---------------------
    # TOP PRODUCTS with PROPER profit margin
    # ---------------------
    top_products = []
    
    # Get top selling products
    top_sales = sales_items.values(
        'product__id', 'product__name', 'product__sku'
    ).annotate(
        revenue=Sum(REVENUE_EXPR),
        quantity=Sum('quantity'),
        avg_price=Avg('sale_price')
    ).order_by('-revenue')[:10]
    
    for product_data in top_sales:
        product_id = product_data['product__id']
        revenue = product_data['revenue'] or Decimal('0')
        quantity = product_data['quantity'] or 0
        avg_price = product_data['avg_price'] or Decimal('0')
        
        # Calculate COGS for this product
        product_cogs = Decimal('0')
        
        if revenue > 0:
            if product_id in product_cogs_map:
                # Use calculated COGS from our map
                product_cogs = Decimal(str(product_cogs_map[product_id]['cogs']))
            else:
                # Calculate using average cost
                avg_cost_result = PurchaseOrderItem.objects.filter(
                    product_id=product_id
                ).aggregate(
                    avg_cost=Avg('unit_cost')
                )
                avg_cost = avg_cost_result['avg_cost'] or Decimal('0')
                product_cogs = avg_cost * quantity
        
        # If still no COGS, use estimated
        if product_cogs == 0 and revenue > 0:
            product_cogs = revenue * Decimal('0.60')
        
        estimated_profit = revenue - product_cogs
        profit_margin = (estimated_profit / revenue * 100) if revenue else Decimal('0')
        
        top_products.append({
            'product__id': product_id,
            'product__name': product_data['product__name'],
            'product__sku': product_data['product__sku'],
            'total_revenue': float(revenue),
            'total_quantity': quantity,
            'avg_price': float(avg_price),
            'estimated_cogs': float(product_cogs),
            'estimated_profit': float(estimated_profit),
            'profit_margin': float(profit_margin)
        })

    # ---------------------
    # NET PROFIT CALCULATION
    # For now, use gross profit as net profit (before operating expenses)
    # ---------------------
    net_profit = gross_profit
    net_profit_margin = gross_margin

    # ---------------------
    # ACCOUNTS RECEIVABLE
    # ---------------------
    accounts_receivable = None
    if store_id:
        store_filter = {'store_id': store_id}
    else:
        store_filter = {}
    
    try:
        accounts_receivable = calculate_accounts_receivable(store_filter)
    except Exception as e:
        print(f"Error calculating accounts receivable: {e}")
        accounts_receivable = {
            'total_ar': Decimal('0'),
            'aging_buckets': {
                'current': {'amount': Decimal('0'), 'count': 0},
                '1_30': {'amount': Decimal('0'), 'count': 0},
                '31_60': {'amount': Decimal('0'), 'count': 0},
                '60_plus': {'amount': Decimal('0'), 'count': 0},
            },
            'invoices': []
        }

    # ---------------------
    # PREPARE CONTEXT
    # ---------------------
    context = {
        'total_revenue': total_revenue,
        'total_cogs': total_cogs,
        'gross_profit': gross_profit,
        'gross_margin': gross_margin,
        'net_profit': net_profit,
        'net_profit_margin': net_profit_margin,
        'cogs_ratio': cogs_ratio,
        
        # Data for tables and charts
        'monthly_trends': monthly_trends_dict,
        'category_performance': category_performance,
        'top_products': top_products,
        
        # Accounts receivable data
        'accounts_receivable': accounts_receivable,
        
        # Filter data
        'stores': StoreLocation.objects.filter(is_active=True),
        'selected_store': store_id,
        'start_date': start_date,
        'end_date': end_date,
        'date_range': f"{start_date.strftime('%Y-%m-%d')} - {end_date.strftime('%Y-%m-%d')}",
        'report_id': f"FIN-{timezone.now():%Y%m%d}-{random.randint(1000,9999)}",
        
        # COGS details for the template
        'cogs_details': {
            'cogs_ratio': float(cogs_ratio),
            'total_cogs': float(total_cogs),
            'gross_margin': float(gross_margin)
        }
    }
    
    return render(request, 'reports/financial_details.html', context)

def calculate_accounts_receivable(store_filter):
    """Calculate accounts receivable"""
    # Get sales with balance > 0 (using balance field from your Sales model)
    ar_sales = Sales.objects.filter(
        balance__gt=0,
        **store_filter
    ).select_related('customer', 'store').order_by('sale_date')
    
    total_ar = ar_sales.aggregate(total=Sum('balance'))['total'] or Decimal('0')
    
    # Categorize by aging
    today = timezone.now().date()
    aging_buckets = {
        'current': {'amount': Decimal('0'), 'count': 0},
        '1_30': {'amount': Decimal('0'), 'count': 0},
        '31_60': {'amount': Decimal('0'), 'count': 0},
        '60_plus': {'amount': Decimal('0'), 'count': 0},
    }
    
    invoices = []
    for sale in ar_sales:
        days_old = (today - sale.sale_date).days
        
        if days_old <= 0:
            bucket = 'current'
        elif days_old <= 30:
            bucket = '1_30'
        elif days_old <= 60:
            bucket = '31_60'
        else:
            bucket = '60_plus'
        
        aging_buckets[bucket]['amount'] += sale.balance
        aging_buckets[bucket]['count'] += 1
        
        # Calculate due date (assuming 30 days credit)
        due_date = sale.sale_date + timedelta(days=30)
        
        invoices.append({
            'id': sale.id,
            'receipt_no': sale.receipt_no,
            'customer': sale.customer.name if sale.customer else 'Walk-in',
            'sale_date': sale.sale_date,
            'due_date': due_date,
            'amount': sale.balance,
            'days_old': days_old,
        })
    
    return {
        'total_ar': total_ar,
        'aging_buckets': aging_buckets,
        'invoices': invoices,
    }
    

@login_required
def export_financial_report(request, format):
    end_date = timezone.now().date()
    start_date = end_date - timedelta(days=90)

    sales_items = SalesItem.objects.filter(
        order__sale_date__range=[start_date, end_date]
    )

    total_revenue = sales_items.aggregate(
        total=Sum(REVENUE_EXPR)
    )['total'] or Decimal('0')

    total_cogs = PurchaseOrderItem.objects.filter(
        order__purchase_date__range=[start_date, end_date]
    ).aggregate(
        total=Sum(COST_EXPR)
    )['total'] or Decimal('0')

    gross_profit = total_revenue - total_cogs


# ============================================================================
# PRODUCT MASTER REPORTS VIEWS
# ============================================================================
# Product Catalog, Category Analysis, Active/Inactive Products,
# SKU/Barcode Listing, Product Creation History
# ============================================================================

@login_required
def productmaster_details(request):
    """Product Master Reports - REAL production view using EXISTING models"""
    
    # Get all data using your existing models
    all_products = Product.objects.select_related('category').prefetch_related('inventories').all()
    
    # 1. Basic Statistics
    total_products = all_products.count()
    active_products = all_products.filter(is_active=True).count()
    inactive_products = total_products - active_products
    
    categories = Category.objects.annotate(product_count=Count('products')).order_by('-product_count')
    categories_count = categories.count()
    
    # Calculate products with SKU
    total_skus = all_products.filter(sku__isnull=False).exclude(sku='').count()
    
    # 2. Stock analysis
    products_with_stock = sum(1 for p in all_products if p.total_stock > 0)
    out_of_stock_products = total_products - products_with_stock
    total_stock = sum(p.total_stock for p in all_products)
    
    # 3. Percentages
    if total_products > 0:
        active_percentage = (active_products / total_products) * 100
        inactive_percentage = (inactive_products / total_products) * 100
        with_stock_percentage = (products_with_stock / total_products) * 100
        out_of_stock_percentage = (out_of_stock_products / total_products) * 100
    else:
        active_percentage = inactive_percentage = with_stock_percentage = out_of_stock_percentage = 0
    
    # 4. Category analysis
    largest_category = categories.first() if categories else None
    smallest_category = categories.last() if categories else None
    empty_categories = categories.filter(product_count=0).count()
    
    # Calculate low stock categories - FIXED SYNTAX
    low_stock_categories = 0
    for category in categories:
        category_products = category.products.all()
        low_stock_count = 0
        for p in category_products:
            if p.inventories.exists():
                inventory = p.inventories.first()
                if p.total_stock <= inventory.reorder_level:
                    low_stock_count += 1
            else:
                # If no inventory record, consider it out of stock
                low_stock_count += 1
        
        if low_stock_count > 0:
            low_stock_categories += 1
    
    # Average products per category
    avg_products_per_category = categories.aggregate(avg=Avg('product_count'))['avg'] or 0
    
    # 5. Paginate products for catalog tab
    page = request.GET.get('page', 1)
    paginator = Paginator(all_products.order_by('name'), 12)  # 12 products per page
    products_page = paginator.get_page(page)
    
    # 6. Products for status tab
    status_page = request.GET.get('status_page', 1)
    status_paginator = Paginator(all_products.order_by('-is_active', 'name'), 20)
    status_products = status_paginator.get_page(status_page)
    
    # 7. Recent products (for timeline)
    recent_products = all_products.order_by('-created_at')[:10]
    
    # 8. SKU/Barcode data
    sku_products = all_products.filter(sku__isnull=False).exclude(sku='')[:50]
    sku_count = all_products.filter(sku__isnull=False).exclude(sku='').count()
    no_sku_count = total_products - sku_count
    barcode_count = all_products.filter(barcode__isnull=False).exclude(barcode='').count()
    no_barcode_count = total_products - barcode_count
    
    # 9. Creation statistics
    today = timezone.now().date()
    month_start = today.replace(day=1)
    week_start = today - timedelta(days=today.weekday())
    
    products_today = all_products.filter(created_at__date=today).count()
    products_this_week = all_products.filter(created_at__date__gte=week_start).count()
    products_this_month = all_products.filter(created_at__date__gte=month_start).count()
    
    # Calculate creation trend
    last_month_start = (month_start - timedelta(days=1)).replace(day=1)
    products_last_month = all_products.filter(
        created_at__date__gte=last_month_start,
        created_at__date__lt=month_start
    ).count()
    
    if products_last_month > 0:
        creation_trend = ((products_this_month - products_last_month) / products_last_month) * 100
    else:
        creation_trend = 0 if products_this_month == 0 else 100
    
    # Most active category (by product count)
    most_active_category = categories.first()
    
    # Average creation rate per month
    if all_products.exists():
        first_product_date = all_products.order_by('created_at').first().created_at.date()
        months_diff = (today.year - first_product_date.year) * 12 + (today.month - first_product_date.month) + 1
        avg_creation_rate = total_products / max(months_diff, 1)
    else:
        avg_creation_rate = 0
    
    # Most active month
    from django.db.models.functions import TruncMonth
    month_counts = all_products.annotate(
        month=TruncMonth('created_at')
    ).values('month').annotate(
        count=Count('id')
    ).order_by('-count')
    
    most_active_month = month_counts.first()['month'].strftime('%B %Y') if month_counts else 'N/A'
    
    context = {
        'current_date': timezone.now(),
        
        # Statistics
        'total_products': total_products,
        'active_products': active_products,
        'inactive_products': inactive_products,
        'categories_count': categories_count,
        'total_skus': total_skus,
        'total_stock': total_stock,
        
        # Stock analysis
        'products_with_stock': products_with_stock,
        'out_of_stock_products': out_of_stock_products,
        
        # Percentages
        'active_percentage': active_percentage,
        'inactive_percentage': inactive_percentage,
        'with_stock_percentage': with_stock_percentage,
        'out_of_stock_percentage': out_of_stock_percentage,
        
        # Category data
        'categories': categories,
        'largest_category': largest_category,
        'smallest_category': smallest_category,
        'empty_categories': empty_categories,
        'low_stock_categories': low_stock_categories,
        'avg_products_per_category': avg_products_per_category,
        'most_active_category': most_active_category,
        
        # Paginated data
        'products': products_page,
        'status_products': status_products,
        'recent_products': recent_products,
        'sku_products': sku_products,
        
        # SKU/Barcode stats
        'sku_count': sku_count,
        'no_sku_count': no_sku_count,
        'barcode_count': barcode_count,
        'no_barcode_count': no_barcode_count,
        
        # Creation stats
        'products_today': products_today,
        'products_this_week': products_this_week,
        'products_this_month': products_this_month,
        'creation_trend': creation_trend,
        'avg_creation_rate': avg_creation_rate,
        'most_active_month': most_active_month,
    }
    
    return render(request, 'reports/productmaster_details.html', context)


@login_required
def export_product_report(request, format):
    """Export product reports in various formats"""
    from django.http import HttpResponse
    import csv
    from reportlab.pdfgen import canvas
    from reportlab.lib.pagesizes import letter
    import io
    
    if format == 'csv':
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename="product_report.csv"'
        
        writer = csv.writer(response)
        writer.writerow(['Product Name', 'SKU', 'Category', 'Status', 'Stock', 'Price'])
        
        for product in Product.objects.all():
            writer.writerow([
                product.name,
                product.sku,
                product.category.name if product.category else '',
                'Active' if product.is_active else 'Inactive',
                product.total_stock,
                product.default_price or 0
            ])
        
        return response
    
    elif format == 'pdf':
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = 'attachment; filename="product_report.pdf"'
        
        buffer = io.BytesIO()
        p = canvas.Canvas(buffer, pagesize=letter)
        p.drawString(100, 750, "Product Master Report")
        p.drawString(100, 730, f"Generated: {timezone.now().strftime('%Y-%m-%d %H:%M')}")
        p.drawString(100, 710, f"Total Products: {Product.objects.count()}")
        p.showPage()
        p.save()
        
        pdf = buffer.getvalue()
        buffer.close()
        response.write(pdf)
        
        return response
    
    return HttpResponse('Invalid format', status=400)


@login_required
def get_product_catalog_data(request):
    """API endpoint for product catalog data (for AJAX)"""
    import json
    from django.core import serializers
    
    products = Product.objects.all()
    data = serializers.serialize('json', products)
    return HttpResponse(data, content_type='application/json')


@login_required
def get_product_statistics(request):
    """API endpoint for product statistics"""
    from django.http import JsonResponse
    
    today = timezone.now().date()
    month_start = today.replace(day=1)
    
    stats = {
        'total_products': Product.objects.count(),
        'active_products': Product.objects.filter(is_active=True).count(),
        'new_this_month': Product.objects.filter(created_at__gte=month_start).count(),
        'out_of_stock': Product.objects.filter(inventories__quantity_in_stock=0).distinct().count(),
        'low_stock': Product.objects.filter(
            inventories__quantity_in_stock__lte=F('inventories__reorder_level')
        ).distinct().count(),
    }
    
    return JsonResponse(stats)


# ============================================================================
# REORDER & LOW STOCK REPORTS VIEWS
# ============================================================================
# Below Reorder Level Alerts, Out-of-stock Report, Store-specific Low Stock,
# Automatic Reorder Suggestions
# ============================================================================

@login_required
def reorder_details(request):
    """Reorder & Low Stock Reports view"""
    
    # Get all active stores
    stores = StoreLocation.objects.filter(is_active=True)
    
    # Get current date for calculations
    today = timezone.now().date()
    
    # 1. BELOW REORDER LEVEL ITEMS
    below_reorder_items = []
    for inventory in Inventory.objects.filter(
        quantity_in_stock__gt=0  # Only items with some stock
    ).select_related('product', 'store'):
        
        # Calculate stock percentage relative to reorder level
        if inventory.reorder_level > 0:
            stock_percentage = (inventory.quantity_in_stock / inventory.reorder_level) * 100
        else:
            stock_percentage = 0
            
        # Determine priority
        if inventory.quantity_in_stock == 0:
            priority = 'critical'
        elif stock_percentage <= 50:
            priority = 'critical'
        elif stock_percentage <= 75:
            priority = 'high'
        elif stock_percentage <= 90:
            priority = 'medium'
        else:
            priority = 'low'
        
        # Calculate average daily sales (last 30 days)
        thirty_days_ago = today - timedelta(days=30)
        total_sales = SalesItem.objects.filter(
            product=inventory.product,
            order__sale_date__gte=thirty_days_ago,
            order__store=inventory.store
        ).aggregate(total_quantity=Sum('quantity'))['total_quantity'] or 0
        
        avg_daily_sales = total_sales / 30 if total_sales > 0 else 0
        
        # Calculate days until stockout
        days_until_stockout = inventory.quantity_in_stock / avg_daily_sales if avg_daily_sales > 0 else 999
        
        # Get days below reorder level (simplified - would need historical tracking)
        days_below_reorder = 0  # You would need to track this historically
        
        if inventory.quantity_in_stock <= inventory.reorder_level:
            below_reorder_items.append({
                'inventory': inventory,
                'stock_percentage': stock_percentage,
                'priority': priority,
                'avg_daily_sales': round(avg_daily_sales, 1),
                'days_until_stockout': round(days_until_stockout, 1),
                'days_below_reorder': days_below_reorder,
                'stock_value': inventory.quantity_in_stock * inventory.product.default_price
            })
    
    # Sort by priority (critical first)
    below_reorder_items.sort(key=lambda x: {
        'critical': 0, 'high': 1, 'medium': 2, 'low': 3
    }[x['priority']])
    
    # 2. OUT OF STOCK ITEMS
    out_of_stock_items = []
    for inventory in Inventory.objects.filter(
        quantity_in_stock=0
    ).select_related('product', 'store'):
        
        # Get last sale date
        last_sale = SalesItem.objects.filter(
            product=inventory.product,
            order__store=inventory.store
        ).order_by('-order__sale_date').first()
        
        # Count backorders/requests (would need a backorder model)
        backorders = 0
        
        # Get supplier info (from purchase orders)
        last_purchase = PurchaseOrderItem.objects.filter(
            product=inventory.product
        ).order_by('-order__purchase_date').first()
        
        supplier = last_purchase.order.supplier if last_purchase else None
        
        # Calculate days out of stock
        days_out = 0  # You would need to track when it went out of stock
        
        # Determine urgency based on sales velocity
        thirty_days_ago = today - timedelta(days=30)
        sales_last_month = SalesItem.objects.filter(
            product=inventory.product,
            order__sale_date__gte=thirty_days_ago,
            order__store=inventory.store
        ).aggregate(total_quantity=Sum('quantity'))['total_quantity'] or 0
        
        if sales_last_month > 20:
            urgency = 'critical'
        elif sales_last_month > 10:
            urgency = 'high'
        elif sales_last_month > 0:
            urgency = 'medium'
        else:
            urgency = 'low'
        
        out_of_stock_items.append({
            'inventory': inventory,
            'last_sale': last_sale.order.sale_date if last_sale else None,
            'backorders': backorders,
            'supplier': supplier,
            'days_out': days_out,
            'urgency': urgency,
            'sales_last_month': sales_last_month
        })
    
    # Sort by urgency
    out_of_stock_items.sort(key=lambda x: {
        'critical': 0, 'high': 1, 'medium': 2, 'low': 3
    }[x['urgency']])
    
    # 3. STORE-SPECIFIC LOW STOCK
    store_specific_data = {}
    for store in stores:
        low_stock_in_store = []
        
        for inventory in Inventory.objects.filter(
            store=store,
            quantity_in_stock__gt=0,  # Has some stock
            quantity_in_stock__lte=F('reorder_level') * 2  # Below 2x reorder level
        ).select_related('product'):
            
            # Get stock in other stores
            other_store_stock = []
            for other_store in stores.exclude(id=store.id):
                try:
                    other_inv = Inventory.objects.get(
                        product=inventory.product,
                        store=other_store
                    )
                    if other_inv.quantity_in_stock > inventory.reorder_level * 1.5:  # Has excess stock
                        other_store_stock.append({
                            'store': other_store,
                            'quantity': other_inv.quantity_in_stock
                        })
                except Inventory.DoesNotExist:
                    continue
            
            # Determine transfer recommendation
            if other_store_stock:
                # Can transfer from other stores
                best_source = max(other_store_stock, key=lambda x: x['quantity'])
                
                # Get the reorder level for the source store
                try:
                    source_inventory = Inventory.objects.get(
                        product=inventory.product,
                        store=best_source['store']
                    )
                    source_reorder_level = source_inventory.reorder_level
                except Inventory.DoesNotExist:
                    source_reorder_level = 10  # Default
                
                # Calculate suggested units
                suggested_units = min(
                    inventory.reorder_level - inventory.quantity_in_stock,
                    best_source['quantity'] - source_reorder_level
                )
                suggested_units = max(suggested_units, 1)  # At least 1 unit
                
                recommendation = {
                    'type': 'transfer',
                    'from_store': best_source['store'],
                    'suggested_units': suggested_units,
                    'lead_time': '1-2 days'
                }
            else:
                # Need to reorder from supplier
                recommendation = {
                    'type': 'reorder',
                    'lead_time': '4-5 days'
                }
            
            low_stock_in_store.append({
                'inventory': inventory,
                'other_store_stock': other_store_stock,
                'recommendation': recommendation
            })
        
        store_specific_data[store] = low_stock_in_store
    
    # 4. SUMMARY STATISTICS
    total_out_of_stock = Inventory.objects.filter(quantity_in_stock=0).count()
    total_below_reorder = len(below_reorder_items)
    
    # Calculate total value at risk
    total_value_at_risk = sum(
        item['stock_value'] for item in below_reorder_items
    ) + sum(
        item['inventory'].reorder_level * item['inventory'].product.default_price
        for item in below_reorder_items if item['inventory'].quantity_in_stock == 0
    )
    
    # Store-wise counts
    store_stats = {}
    for store in stores:
        store_stats[store.name] = {
            'low_stock': Inventory.objects.filter(
                store=store,
                quantity_in_stock__gt=0,
                quantity_in_stock__lte=F('reorder_level')
            ).count(),
            'out_of_stock': Inventory.objects.filter(
                store=store,
                quantity_in_stock=0
            ).count()
        }
    
    context = {
        'below_reorder_items': below_reorder_items[:50],  # Limit to 50 items
        'out_of_stock_items': out_of_stock_items[:50],
        'store_specific_data': store_specific_data,
        'stores': stores,
        'total_out_of_stock': total_out_of_stock,
        'total_below_reorder': total_below_reorder,
        'total_value_at_risk': total_value_at_risk,
        'store_stats': store_stats,
        'report_id': f"STOCK-ALERT-{today.strftime('%Y%m%d')}-001",
        'report_date': today,
        'urgent_count': len([item for item in below_reorder_items if item['priority'] == 'critical']),
        'warning_count': len([item for item in below_reorder_items if item['priority'] == 'high']),
        'monitor_count': len([item for item in below_reorder_items if item['priority'] in ['medium', 'low']]),
    }
    
    return render(request, 'reports/reorder_details.html', context)


# ============================================================================
# STORE LOCATION REPORTS VIEWS
# ============================================================================
# Store Capacity/Utilization, Store Activity Analysis, Branch-wise Store Comparison,
# Default Store Analysis
# ============================================================================

def decimal_to_float(obj):
    """Convert Decimal objects to float for JSON serialization"""
    if isinstance(obj, Decimal):
        return float(obj)
    raise TypeError(f"Object of type {obj.__class__.__name__} is not JSON serializable")

@login_required
def stocklocation_details(request):
    """Store Location Reports dashboard view - Fully Dynamic"""
    
    # Get all active stores
    stores = StoreLocation.objects.filter(is_active=True)
    
    # Date range - default to current month
    start_date_str = request.GET.get('start_date')
    end_date_str = request.GET.get('end_date')
    
    if start_date_str and end_date_str:
        try:
            start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
            end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()
        except ValueError:
            start_date = date.today().replace(day=1)
            end_date = date.today()
    else:
        start_date = date.today().replace(day=1)
        end_date = date.today()
    
    # Get actual sales status from model
    actual_sales_statuses = Sales.objects.values_list('status', flat=True).distinct()
    
    # Determine which status to use for "completed" sales
    possible_statuses = ['FULFILLED', 'Fulfilled', 'COMPLETED', 'Completed']
    sales_status = None
    
    for status in possible_statuses:
        if status in actual_sales_statuses:
            sales_status = status
            break
    
    # If no matching status found, use the first available status
    if not sales_status and actual_sales_statuses:
        sales_status = actual_sales_statuses[0]
    
    # Calculate store metrics
    store_reports = []
    
    # Get all store sales for calculating averages
    all_store_sales = {}
    for store in stores:
        sales_data = Sales.objects.filter(
            store=store,
            sale_date__range=[start_date, end_date],
            status=sales_status
        ).aggregate(
            total_sales=Sum('total_amount'),
            total_transactions=Count('id'),
            avg_sale_value=Avg('total_amount')
        )
        all_store_sales[store.id] = sales_data['total_sales'] or 0
    
    # Calculate average sales across all stores
    total_sales_all_stores = sum(all_store_sales.values())
    avg_sales_all_stores = float(total_sales_all_stores) / len(stores) if stores else 0 
    
    for store in stores:
        # Sales metrics for the period
        sales_data = Sales.objects.filter(
            store=store,
            sale_date__range=[start_date, end_date],
            status=sales_status
        ).aggregate(
            total_sales=Sum('total_amount'),
            total_transactions=Count('id'),
            avg_sale_value=Avg('total_amount')
        )
        
        # Convert Decimal to float for calculations
        total_sales_val = float(sales_data['total_sales'] or 0)
        avg_sale_val = float(sales_data['avg_sale_value'] or 0)
        
        # Inventory metrics
        inventory_data = Inventory.objects.filter(store=store).aggregate(
            total_skus=Count('product', distinct=True),
            total_items=Sum('quantity_in_stock'),
            low_stock_items=Count('id', filter=Q(quantity_in_stock__lte=F('reorder_level'))),
            zero_stock_items=Count('id', filter=Q(quantity_in_stock=0))
        )
        
        # Get purchase order status from model
        purchase_statuses = PurchaseOrder.objects.values_list('status', flat=True).distinct()
        purchase_status = 'received' if 'received' in purchase_statuses else purchase_statuses[0] if purchase_statuses else None
        
        # Purchase metrics
        purchase_data = {}
        if purchase_status:
            purchase_data = PurchaseOrder.objects.filter(
                store=store,
                purchase_date__range=[start_date, end_date],
                status=purchase_status
            ).aggregate(
                total_purchases=Sum('total_cost'),
                purchase_orders=Count('id')
            )
            # Convert Decimal to float
            if purchase_data['total_purchases']:
                purchase_data['total_purchases'] = float(purchase_data['total_purchases'])
        else:
            purchase_data = {'total_purchases': 0, 'purchase_orders': 0}
        
        # Get stock transfer status from model
        transfer_statuses = StockTransfer.objects.values_list('status', flat=True).distinct()
        completed_status = 'completed' if 'completed' in transfer_statuses else transfer_statuses[0] if transfer_statuses else None
        
        # Transfer metrics
        transfers_out = {'total_items': 0}
        transfers_out_value = 0
        
        if completed_status:
            transfers_out = StockTransfer.objects.filter(
                from_store=store,
                transfer_date__range=[start_date, end_date],
                status=completed_status
            ).aggregate(
                total_items=Sum('items__quantity')
            )
            
            completed_transfers = StockTransfer.objects.filter(
                from_store=store,
                transfer_date__range=[start_date, end_date],
                status=completed_status
            )
            for transfer in completed_transfers:
                transfers_out_value += float(transfer.total_value or 0)
        
        transfers_in = {'total_items': 0}
        if completed_status:
            transfers_in = StockTransfer.objects.filter(
                to_store=store,
                transfer_date__range=[start_date, end_date],
                status=completed_status
            ).aggregate(
                total_items=Sum('items__quantity')
            )
        
        # Stock movement activity (last 7 days)
        week_ago = end_date - timedelta(days=7)
        recent_activity = StockMovement.objects.filter(
            store=store,
            timestamp__date__gte=week_ago
        ).count()
        
        # Capacity/utilization calculation
        total_batches = InventoryBatch.objects.filter(store=store, remaining_quantity__gt=0)
        total_units = total_batches.aggregate(total=Sum('remaining_quantity'))['total'] or 0
        
        max_quantity_ever = InventoryBatch.objects.filter(store=store).aggregate(
            max_quantity=Sum('quantity')
        )['max_quantity'] or 0
        
        if max_quantity_ever > 0:
            utilization_percentage = min(100, round((total_units / max_quantity_ever) * 100, 1))
        else:
            utilization_percentage = 0
        
        # Determine performance score (0-10)
        performance_score = 6.0
        
        # Adjust based on sales performance (0-2 points)
        store_sales = total_sales_val
        if avg_sales_all_stores > 0:
            sales_ratio = store_sales / avg_sales_all_stores
            if sales_ratio >= 1.5:
                performance_score += 2.0
            elif sales_ratio >= 1.2:
                performance_score += 1.5
            elif sales_ratio >= 0.8:
                performance_score += 1.0
            elif sales_ratio >= 0.5:
                performance_score += 0.5
        
        # Adjust based on inventory utilization (0-1.5 points)
        if 70 <= utilization_percentage <= 85:
            performance_score += 1.5
        elif 60 <= utilization_percentage < 70 or 85 < utilization_percentage <= 90:
            performance_score += 1.0
        elif 50 <= utilization_percentage < 60 or 90 < utilization_percentage <= 95:
            performance_score += 0.5
        elif utilization_percentage > 95:
            performance_score -= 0.5
        
        # Adjust based on stock availability (0-1 point)
        total_skus = inventory_data['total_skus'] or 0
        low_stock_items = inventory_data['low_stock_items'] or 0
        if total_skus > 0:
            low_stock_ratio = low_stock_items / total_skus
            if low_stock_ratio <= 0.1:
                performance_score += 1.0
            elif low_stock_ratio <= 0.2:
                performance_score += 0.5
            elif low_stock_ratio > 0.3:
                performance_score -= 0.5
        
        # Adjust based on activity (0-0.5 points)
        if recent_activity > 20:
            performance_score += 0.5
        elif recent_activity > 10:
            performance_score += 0.25
        
        # Cap score between 0 and 10
        performance_score = max(0, min(10, round(performance_score, 1)))
        
        # Determine performance category
        if performance_score >= 8.5:
            performance_category = 'excellent'
            performance_badge = 'performance-excellent'
            score_class = 'score-excellent'
        elif performance_score >= 7.0:
            performance_category = 'good'
            performance_badge = 'performance-good'
            score_class = 'score-good'
        else:
            performance_category = 'fair'
            performance_badge = 'performance-fair'
            score_class = 'score-fair'
        
        # Calculate efficiency score based on multiple factors
        efficiency_score = 70
        
        # Add based on utilization (max 10 points)
        if 70 <= utilization_percentage <= 85:
            efficiency_score += 10
        elif 60 <= utilization_percentage < 70 or 85 < utilization_percentage <= 90:
            efficiency_score += 5
        
        # Add based on sales performance (max 10 points)
        if store_sales > 0 and avg_sales_all_stores > 0:
            sales_eff = min(store_sales / avg_sales_all_stores * 10, 10)
            efficiency_score += sales_eff
        
        # Add based on low stock management (max 5 points)
        if total_skus > 0:
            low_stock_ratio = low_stock_items / total_skus
            if low_stock_ratio <= 0.1:
                efficiency_score += 5
            elif low_stock_ratio <= 0.2:
                efficiency_score += 3
        
        # Cap efficiency score at 100%
        efficiency_score = min(100, efficiency_score)
        
        # Calculate growth rate
        previous_month_start = (start_date - timedelta(days=30)).replace(day=1)
        previous_month_end = start_date - timedelta(days=1)
        
        previous_sales = Sales.objects.filter(
            store=store,
            sale_date__range=[previous_month_start, previous_month_end],
            status=sales_status
        ).aggregate(total_sales=Sum('total_amount'))['total_sales'] or 0
        previous_sales = float(previous_sales)

        current_sales = store_sales
        if previous_sales > 0:
            growth_rate = ((current_sales - previous_sales) / previous_sales) * 100
            # Cap unrealistic growth percentages
            if growth_rate > 500:
                growth_rate = 500
        elif current_sales > 0:
            growth_rate = 100  # First time sales
        else:
            growth_rate = 0
            
        
        # Generate dynamic color based on store ID
        color_index = store.id % 6
        color_classes = ['store-primary', 'store-secondary', 'store-success', 
                        'store-warning', 'store-danger', 'store-info']
        store_color_class = color_classes[color_index]
        
        # Build store report with float values for JSON serialization
        store_report = {
            'store': store,
            'color_class': store_color_class,
            'sales_data': {
                'total_sales': total_sales_val,
                'total_transactions': sales_data['total_transactions'] or 0,
                'avg_sale_value': avg_sale_val,
            },
            'inventory_data': {
                'total_skus': inventory_data['total_skus'] or 0,
                'total_items': inventory_data['total_items'] or 0,
                'low_stock_items': inventory_data['low_stock_items'] or 0,
                'zero_stock_items': inventory_data['zero_stock_items'] or 0,
            },
            'purchase_data': purchase_data,
            'transfers_out': {
                'total_items': transfers_out['total_items'] or 0,
                'total_value': transfers_out_value
            },
            'transfers_in': transfers_in,
            'recent_activity': recent_activity,
            'utilization_percentage': utilization_percentage,
            'performance_score': performance_score,
            'performance_category': performance_category,
            'performance_badge': performance_badge,
            'score_class': score_class,
            'efficiency_score': round(efficiency_score),
            'growth_rate': round(growth_rate, 1),
            'total_units': total_units,
            'max_capacity': max_quantity_ever,
            'is_default': store.is_default,
        }
        
        store_reports.append(store_report)
    
    # Get default store
    default_store = stores.filter(is_default=True).first()
    
    # Get default store report
    default_store_report = None
    if default_store:
        for report in store_reports:
            if report['store'].id == default_store.id:
                default_store_report = report
                break
    
    # Generate chart data
    # Sales trend chart data
    sales_trend_labels = []
    sales_trend_data = {}
    
    # Generate labels for the last 7 days
    for i in range(6, -1, -1):
        day = end_date - timedelta(days=i)
        sales_trend_labels.append(day.strftime('%b %d'))
    
    # Get sales data for each store for the last 7 days
    for store_report in store_reports:
        store = store_report['store']
        store_data = []
        
        for i in range(6, -1, -1):
            day = end_date - timedelta(days=i)
            daily_sales = Sales.objects.filter(
                store=store,
                sale_date=day,
                status=sales_status
            ).aggregate(total=Sum('total_amount'))['total'] or 0
            store_data.append(float(daily_sales) / 1000000)  # Convert to millions
            
        sales_trend_data[store.name] = {
            'data': store_data,
            'color': get_dynamic_chart_color(store.id)
        }
    
    # Capacity utilization chart data
    capacity_labels = [report['store'].name for report in store_reports]
    capacity_data = [report['utilization_percentage'] for report in store_reports]
    capacity_colors = [get_dynamic_chart_color(report['store'].id) for report in store_reports]
    
    # Activity chart data (transactions per day for last 7 days - not hourly)
    activity_labels = []
    activity_data = {}
    
    # Generate labels for the last 7 days
    for i in range(6, -1, -1):
        day = end_date - timedelta(days=i)
        activity_labels.append(day.strftime('%a %d'))
    
    # Generate activity data based on recent transactions
    for store_report in store_reports[:3]:  # Show only first 3 stores for clarity
        store = store_report['store']
        
        daily_data = []
        
        for i in range(6, -1, -1):
            day = end_date - timedelta(days=i)
            
            # Count transactions for this day
            daily_transactions = Sales.objects.filter(
                store=store,
                sale_date=day,
                status=sales_status
            ).count()
            
            daily_data.append(daily_transactions)
        
        activity_data[store.name] = {
            'data': daily_data,
            'color': get_dynamic_chart_color(store.id)
        }
    
    # Comparison radar chart data
    comparison_labels = ['Sales', 'Efficiency', 'Growth', 'Activity', 'Utilization', 'Stock Health']
    comparison_data = {}
    
    for store_report in store_reports[:3]:  # Show only first 3 stores
        store = store_report['store']
        
        # Calculate radar scores (0-100)
        sales_score = min(100, (store_report['sales_data']['total_sales'] or 0) / 1000000 * 20)
        efficiency_score = store_report['efficiency_score']
        growth_score = min(100, max(0, 50 + store_report['growth_rate']))
        
        # Activity score based on recent activity
        activity_score = min(100, store_report['recent_activity'] * 5)
        
        # Utilization score
        utilization_score = store_report['utilization_percentage']
        
        # Stock health score (higher is better - less low stock items)
        total_skus = store_report['inventory_data']['total_skus'] or 1
        low_stock_items = store_report['inventory_data']['low_stock_items'] or 0
        stock_health_score = max(0, 100 - (low_stock_items / total_skus * 100))
        
        radar_data = [
            sales_score,
            efficiency_score,
            growth_score,
            activity_score,
            utilization_score,
            stock_health_score
        ]
        
        comparison_data[store.name] = {
            'data': radar_data,
            'color': get_dynamic_chart_color(store.id),
            'border_color': get_dynamic_border_color(store.id)
        }
    
    # Calculate best performing store
    best_performing = None
    if store_reports:
        best_performing_report = max(store_reports, key=lambda x: x['performance_score'])
        best_performing = best_performing_report['store']
    
    # Calculate total sales across all stores
    total_sales = sum((report['sales_data']['total_sales'] or 0) for report in store_reports)
    
    # Calculate average utilization
    avg_utilization = 0
    if store_reports:
        avg_utilization = round(sum(report['utilization_percentage'] for report in store_reports) / len(store_reports), 1)
    
    # Prepare context - using custom JSON encoder for Decimal objects
    context = {
        'stores': stores,
        'store_reports': store_reports,
        'default_store': default_store,
        'default_store_report': default_store_report,
        'start_date': start_date,
        'end_date': end_date,
        'total_stores': stores.count(),
        'total_sales': total_sales,
        'avg_utilization': avg_utilization,
        'best_performing': best_performing,
        'sales_status': sales_status,
        # Chart data as JSON - using custom encoder
        'sales_trend_labels': json.dumps(sales_trend_labels),
        'sales_trend_data': json.dumps(sales_trend_data, default=decimal_to_float),
        'capacity_labels': json.dumps(capacity_labels),
        'capacity_data': json.dumps(capacity_data, default=decimal_to_float),
        'capacity_colors': json.dumps(capacity_colors),
        'activity_labels': json.dumps(activity_labels),
        'activity_data': json.dumps(activity_data, default=decimal_to_float),
        'comparison_labels': json.dumps(comparison_labels),
        'comparison_data': json.dumps(comparison_data, default=decimal_to_float),
    }
    
    return render(request, 'reports/stocklocation_details.html', context)

# Dynamic color helper functions
def get_dynamic_chart_color(store_id):
    """Generate a consistent chart color based on store ID"""
    colors = [
        '#4A90E2',  # Blue
        '#36B9CC',  # Cyan
        '#1CC88A',  # Green
        '#F6C23E',  # Yellow
        '#E74A3B',  # Red
        '#6C757D',  # Gray
        '#4E73DF',  # Indigo
        '#1CC88A',  # Teal
        '#F6C23E',  # Orange
    ]
    return colors[store_id % len(colors)]


def get_dynamic_border_color(store_id):
    """Generate a consistent border color based on store ID"""
    colors = [
        '#357ABD',  # Darker Blue
        '#2A9CA5',  # Darker Cyan
        '#17A673',  # Darker Green
        '#F4B619',  # Darker Yellow
        '#E02D1B',  # Darker Red
        '#495057',  # Darker Gray
        '#2E59D9',  # Darker Indigo
        '#17A673',  # Darker Teal
        '#F4B619',  # Darker Orange
    ]
    return colors[store_id % len(colors)]


# ============================================================================
# PRICING REPORTS VIEWS
# ============================================================================
# Product Pricing List, Unit Conversion Report, Price Comparison,
# Missing Pricing Report
# ============================================================================

@login_required
def productpricing_details(request):  
    """Pricing reports view with detailed analysis"""
    
    today = timezone.now()
    currency = request.GET.get('currency', 'UGX')
    
    # Get filter parameters
    category_filter = request.GET.get('category', '')
    price_range_filter = request.GET.get('price_range', '')
    margin_filter = request.GET.get('margin', '')
    search_query = request.GET.get('search', '')
    
    # Base queryset for products
    products = Product.objects.filter(is_active=True).prefetch_related(
        'unit_prices', 'category'
    )
    
    # Apply filters
    if search_query:
        products = products.filter(
            Q(name__icontains=search_query) |
            Q(sku__icontains=search_query)
        )
    
    if category_filter:
        products = products.filter(category__name=category_filter)
    
    # Prepare product pricing data
    product_pricing_data = []
    total_selling_price = Decimal('0')
    total_cost_price = Decimal('0')
    products_count = 0
    high_margin_count = 0
    low_margin_count = 0
    total_product_value = Decimal('0')
    
    for product in products:
        default_unit_price = product.unit_prices.first()
        
        if default_unit_price:
            selling_price = default_unit_price.price
            
            # Get cost from purchase history or use default
            purchase_items = PurchaseOrderItem.objects.filter(product=product)
            avg_cost = purchase_items.aggregate(avg=Avg('unit_cost'))['avg']
            cost_price = avg_cost or (selling_price * Decimal('0.65'))
            
            # Calculate margin
            if cost_price > 0:
                margin = ((selling_price - cost_price) / cost_price) * 100
            else:
                margin = 0
            
            # Apply filters
            if price_range_filter:
                if price_range_filter == 'low' and selling_price >= 100000:
                    continue
                elif price_range_filter == 'medium' and (selling_price < 100000 or selling_price > 500000):
                    continue
                elif price_range_filter == 'high' and selling_price <= 500000:
                    continue
            
            if margin_filter:
                if margin_filter == 'low' and margin >= 30:
                    continue
                elif margin_filter == 'medium' and (margin < 30 or margin > 50):
                    continue
                elif margin_filter == 'high' and margin <= 50:
                    continue
            
            # Determine price tier
            if margin >= 50:
                tier = 'Premium'
                tier_class = 'tier-premium'
                high_margin_count += 1
            elif margin >= 30:
                tier = 'Standard'
                tier_class = 'tier-standard'
            else:
                tier = 'Economy'
                tier_class = 'tier-economy'
                low_margin_count += 1
            
            # Determine margin badge
            if margin >= 50:
                margin_class = 'margin-high'
            elif margin >= 30:
                margin_class = 'margin-medium'
            else:
                margin_class = 'margin-low'
            
            # Static price change for now
            price_change = Decimal('0')
            if price_change > 0:
                price_change_class = 'price-up'
                price_change_icon = 'trending-up'
                change_text = f"+{price_change:.1f}%"
            elif price_change < 0:
                price_change_class = 'price-down'
                price_change_icon = 'trending-down'
                change_text = f"{price_change:.1f}%"
            else:
                price_change_class = 'price-stable'
                price_change_icon = 'minus'
                change_text = 'Stable'
            
            product_data = {
                'sku': product.sku,
                'name': product.name,
                'category': product.category.name if product.category else 'Uncategorized',
                'cost_price': cost_price,
                'selling_price': selling_price,
                'margin': margin,
                'margin_class': margin_class,
                'tier': tier,
                'tier_class': tier_class,
                'price_change': price_change,
                'price_change_class': price_change_class,
                'price_change_icon': price_change_icon,
                'change_text': change_text,
                'last_change_date': (today - timedelta(days=7)).strftime('%Y-%m-%d'),
                'last_change_reason': 'Market adjustment',
            }
            
            product_pricing_data.append(product_data)
            total_selling_price += selling_price
            total_cost_price += cost_price
            total_product_value += selling_price
            products_count += 1
    
    # Sort products
    sort_by = request.GET.get('sort', 'name')
    if sort_by == 'price_high':
        product_pricing_data.sort(key=lambda x: x['selling_price'], reverse=True)
    elif sort_by == 'price_low':
        product_pricing_data.sort(key=lambda x: x['selling_price'])
    elif sort_by == 'margin':
        product_pricing_data.sort(key=lambda x: x['margin'], reverse=True)
    else:
        product_pricing_data.sort(key=lambda x: x['name'])
    
    # Calculate statistics
    avg_selling_price = total_selling_price / products_count if products_count > 0 else 0
    avg_cost_price = total_cost_price / products_count if products_count > 0 else 0
    avg_margin = ((avg_selling_price - avg_cost_price) / avg_cost_price * 100) if avg_cost_price > 0 else 0
    
    # Get unit conversion data (products with multiple unit prices)
    unit_conversion_data = []
    for product in products:
        unit_prices = product.unit_prices.all()
        if len(unit_prices) >= 2:
            # Sort by conversion factor (smallest to largest)
            sorted_prices = sorted(unit_prices, key=lambda x: x.conversion_factor)
            base_unit = sorted_prices[0]
            pack_unit = sorted_prices[-1]
            
            if pack_unit.conversion_factor > base_unit.conversion_factor:
                pack_price_per_base_unit = pack_unit.price / pack_unit.conversion_factor
                base_price_per_unit = base_unit.price
                savings_pct = ((base_price_per_unit - pack_price_per_base_unit) / base_price_per_unit * 100) if base_price_per_unit > 0 else 0
                
                unit_conversion_data.append({
                    'product': product.name,
                    'sku': product.sku,
                    'base_unit': base_unit.unit.name,
                    'pack_unit': f"{pack_unit.conversion_factor}{base_unit.unit.abbreviation}",
                    'conversion_factor': f"1:{int(pack_unit.conversion_factor)}",
                    'base_price': base_unit.price,
                    'pack_price': pack_unit.price,
                    'unit_price': pack_price_per_base_unit,
                    'savings_pct': savings_pct,
                })
    
    # Get missing pricing data
    missing_pricing_products = Product.objects.filter(
        is_active=True,
        unit_prices__isnull=True
    )
    
    missing_prices_count = missing_pricing_products.count()
    missing_pricing_data = []
    overdue_count = 0
    new_products_count = 0
    total_days_missing = 0
    
    for product in missing_pricing_products[:10]:  # Limit to 10 for display
        days_without_price = (today - product.created_at).days
        total_days_missing += days_without_price
        
        if days_without_price > 7:
            overdue_count += 1
            status = 'Overdue'
            status_class = 'bg-danger'
        else:
            new_products_count += 1
            status = 'New Product'
            status_class = 'bg-warning'
        
        # Find similar products for price suggestion
        similar_products = Product.objects.filter(
            category=product.category,
            unit_prices__isnull=False
        )[:3]
        
        avg_similar_price = 0
        if similar_products.exists():
            avg_prices = []
            for sim_product in similar_products:
                price = sim_product.unit_prices.first()
                if price:
                    avg_prices.append(price.price)
            if avg_prices:
                avg_similar_price = sum(avg_prices) / len(avg_prices)
        
        # Get cost price from purchase history
        purchase_items = PurchaseOrderItem.objects.filter(product=product)
        avg_cost = purchase_items.aggregate(avg=Avg('unit_cost'))['avg'] or Decimal('0')
        
        missing_pricing_data.append({
            'sku': product.sku,
            'name': product.name,
            'category': product.category.name if product.category else 'Uncategorized',
            'date_added': product.created_at.strftime('%Y-%m-%d'),
            'days_without_price': days_without_price,
            'similar_products_count': similar_products.count(),
            'suggested_price': avg_similar_price,
            'cost_price': avg_cost,
            'status': status,
            'status_class': status_class,
        })
    
    # Calculate margin by category
    categories = Category.objects.all()
    category_margins = []
    for category in categories:
        category_products = products.filter(category=category)
        if category_products.exists():
            total_margin = 0
            count = 0
            for product in category_products:
                unit_price = product.unit_prices.first()
                if unit_price:
                    purchase_items = PurchaseOrderItem.objects.filter(product=product)
                    avg_cost = purchase_items.aggregate(avg=Avg('unit_cost'))['avg']
                    cost_price = avg_cost or (unit_price.price * Decimal('0.65'))
                    if cost_price > 0:
                        margin = ((unit_price.price - cost_price) / cost_price) * 100
                        total_margin += margin
                        count += 1
            if count > 0:
                category_margins.append({
                    'name': category.name,
                    'avg_margin': total_margin / count
                })
    
    # Find highest and lowest margin categories
    if category_margins:
        highest = max(category_margins, key=lambda x: x['avg_margin'])
        lowest = min(category_margins, key=lambda x: x['avg_margin'])
        highest_margin_category = highest['name']
        highest_margin_pct = highest['avg_margin']
        lowest_margin_category = lowest['name']
        lowest_margin_pct = lowest['avg_margin']
    else:
        highest_margin_category = 'N/A'
        highest_margin_pct = 0
        lowest_margin_category = 'N/A'
        lowest_margin_pct = 0
    
    # Prepare context
    context = {
        # Report metadata
        'report_id': f"PRICE-{today.strftime('%Y%m%d')}-001",
        'report_date': today.strftime('%B %Y'),
        'currency': currency,
        'products_analyzed': products_count,
        'missing_prices_count': missing_prices_count,
        
        # Overall statistics
        'avg_selling_price': avg_selling_price,
        'avg_cost_price': avg_cost_price,
        'avg_margin': avg_margin,
        'avg_price_change': Decimal('2.5'),  # Static for now
        'price_changes_this_month': 24,  # Static for now
        
        # Additional statistics
        'total_product_value': total_product_value,
        'high_margin_count': high_margin_count,
        'low_margin_count': low_margin_count,
        
        # Margin analysis
        'highest_margin_category': highest_margin_category,
        'highest_margin_pct': highest_margin_pct,
        'lowest_margin_category': lowest_margin_category,
        'lowest_margin_pct': lowest_margin_pct,
        
        # Data for templates
        'product_pricing_data': product_pricing_data,
        'unit_conversion_data': unit_conversion_data,
        'missing_pricing_data': missing_pricing_data,
        
        # Missing pricing statistics
        'new_products_count': new_products_count,
        'overdue_count': overdue_count,
        'avg_days_missing': total_days_missing / missing_prices_count if missing_prices_count > 0 else 0,
        'default_margin': 50,
        
        # Filter options
        'categories': categories,
        'selected_category': category_filter,
        'selected_price_range': price_range_filter,
        'selected_margin': margin_filter,
        'selected_sort': sort_by,
        'search_query': search_query,
    }
    
    return render(request, 'reports/productpricing_details.html', context)
# ============================================================================
# PERFORMANCE REPORTS VIEWS
# ============================================================================
# Stock Turnover Rate, Fast vs Slow Movers, Product Sales vs Purchases,
# Category Performance
# ============================================================================

@login_required
def performace_details(request):  
    """Reports dashboard view"""
    context = {}
    return render(request, 'reports/performace_details.html', context)


# ============================================================================
# LIFECYCLE REPORTS VIEWS
# ============================================================================
# New Products Added, Inactive Products Report, Product Popularity Analysis,
# Seasonal Product Analysis
# ============================================================================

@login_required
def lifecycle_details(request):  
    """Reports dashboard view"""
    context = {}
    return render(request, 'reports/lifecycle_details.html', context)


# ============================================================================
# OPERATIONAL REPORTS VIEWS
# ============================================================================
# User Activity Report, Stock Turnover Analysis, Reorder Point Effectiveness,
# Transfer Impact Analysis
# ============================================================================

@login_required
def operational_details(request):  
    """Reports dashboard view"""
    context = {}
    return render(request, 'reports/operational_details.html', context)


# ============================================================================
# AUDIT & COMPLIANCE REPORTS VIEWS
# ============================================================================
# Complete Stock Movement Audit, Expiry Compliance Report, Batch Traceability Report,
# Product Information Completeness, Price Standardization Check
# ============================================================================

@login_required
def auditcompliance_details(request):  
    """Reports dashboard view"""
    context = {}
    return render(request, 'reports/auditcompliance_details.html', context)


# ============================================================================
# CORRELATION REPORTS VIEWS
# ============================================================================
# Stock Availability vs Sales, Reorder Effectiveness Analysis,
# Stock Accuracy Reconciliation
# ============================================================================

@login_required
def correlation_details(request):  
    """Reports dashboard view"""
    context = {}
    return render(request, 'reports/correlation_details.html', context)






























@login_required
def reports_view(request, report_type):
    """Handle inventory reports with dynamic templates"""
    
    # Get filters from request
    store_id = request.GET.get('store_id')
    product_id = request.GET.get('product_id')
    days = int(request.GET.get('days', 30))
    page = int(request.GET.get('page', 1))
    per_page = int(request.GET.get('per_page', 50))
    
    # Map report types to methods
    report_methods = {
        'stock_summary': {
            'method': InventoryReports.generate_stock_summary_report,
            'title': 'Stock Summary Report',
            'subtitle': 'Complete inventory overview with stock values',
            'show_filters': False,
            'show_date_filters': False,
        },
        'low_stock': {
            'method': InventoryReports.generate_low_stock_alert_report,
            'title': 'Low Stock Alert Report',
            'subtitle': 'Products below reorder level requiring attention',
            'show_filters': False,
            'show_date_filters': False,
        },
        'store_performance': {
            'method': lambda: InventoryReports.generate_store_performance_report(store_id),
            'title': 'Store Performance Report',
            'subtitle': 'Store-wise inventory metrics and performance',
            'show_filters': True,
            'show_date_filters': False,
        },
        'category_analysis': {
            'method': InventoryReports.generate_category_analysis_report,
            'title': 'Category Analysis Report',
            'subtitle': 'Inventory performance by product category',
            'show_filters': False,
            'show_date_filters': False,
        },
        'product_movement': {
            'method': lambda: InventoryReports.generate_product_movement_report(product_id, days),
            'title': f'Product Movement Report ({days} days)',
            'subtitle': 'Sales velocity and stock movement analysis',
            'show_filters': True,
            'show_date_filters': False,
        },
        'abc_analysis': {
            'method': InventoryReports.generate_abc_analysis_report,
            'title': 'ABC Analysis Report',
            'subtitle': 'Inventory classification by value importance',
            'show_filters': False,
            'show_date_filters': False,
        },
        'inventory_valuation': {
            'method': InventoryReports.generate_inventory_valuation_report,
            'title': 'Inventory Valuation Report',
            'subtitle': 'Complete inventory valuation with cost analysis',
            'show_filters': False,
            'show_date_filters': False,
        },
        'stock_transfer': {
            'method': lambda: InventoryReports.generate_stock_transfer_report(store_id),
            'title': 'Stock Transfer Report',
            'subtitle': 'Transfer status and movement tracking',
            'show_filters': True,
            'show_date_filters': False,
        },
        'product_availability': {
            'method': lambda: InventoryReports.generate_product_availability_report(product_id),
            'title': 'Product Availability Report',
            'subtitle': 'Product availability across all stores',
            'show_filters': True,
            'show_date_filters': False,
        },
    }
    
    if report_type not in report_methods:
        return render(request, 'reports/error.html', {
            'error': 'Invalid report type',
            'report_type': report_type
        })
    
    # Get report configuration
    config = report_methods[report_type]
    
    try:
        # Generate report data
        report_data = config['method']()
        
        # Process report data for template
        processed_data = process_report_data(report_data, report_type)
        
        # Add pagination if needed
        if 'data' in processed_data and len(processed_data['data']) > per_page:
            paginator = Paginator(processed_data['data'], per_page)
            page_obj = paginator.get_page(page)
            processed_data['data'] = page_obj
        
        # Prepare context
        context = {
            'report_type': report_type,
            'report_title': config['title'],
            'report_subtitle': config['subtitle'],
            'report_data': processed_data,
            'show_filters': config['show_filters'],
            'show_date_filters': config.get('show_date_filters', False),
            'filters': {
                'store_id': store_id,
                'product_id': product_id,
                'days': days,
                'date_from': request.GET.get('date_from'),
                'date_to': request.GET.get('date_to'),
            },
            'currency_symbol': 'UGX',
        }
        
        # Add filter options
        from app.models import StoreLocation, Product
        if config['show_filters']:
            context['stores'] = StoreLocation.objects.filter(is_active=True)
            context['products'] = Product.objects.filter(is_active=True)
        
        # Check if it's an AJAX request
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return JsonResponse({'success': True, 'data': report_data})
        
        # Regular request - render template
        return render(request, 'reports/report_view.html', context)
        
    except Exception as e:
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return JsonResponse({'success': False, 'error': str(e)})
        
        return render(request, 'reports/report_view.html', {
            'error': str(e),
            'report_type': report_type,
            'report_title': config['title']
        })


def process_report_data(report_data, report_type):
    """Process raw report data for template display"""
    processed = {
        'generated_at': report_data.get('generated_at', timezone.now()),
        'report_type': report_data.get('report_type', report_type),
        'data': [],
        'headers': [],
        'summary_stats': [],
        'chart_data': None,
    }
    
    # Extract summary statistics
    summary_stats = extract_summary_stats(report_data, report_type)
    if summary_stats:
        processed['summary_stats'] = summary_stats
    
    # Process main data
    if 'data' in report_data and report_data['data']:
        if isinstance(report_data['data'], list) and len(report_data['data']) > 0:
            # Get headers from first item
            first_item = report_data['data'][0]
            if isinstance(first_item, dict):
                processed['headers'] = list(first_item.keys())
                processed['data'] = report_data['data']
    
    # Add category summary if available
    if 'category_summary' in report_data:
        processed['category_summary'] = report_data['category_summary']
    
    # Add insights if available
    if report_data.get('total_products'):
        processed['insights'] = generate_insights(report_data, report_type)
    
    # Generate chart data if needed
    if len(processed['data']) > 0:
        processed['chart_data'] = generate_chart_data(processed['data'], report_type)
    
    return processed


def extract_summary_stats(report_data, report_type):
    """Extract summary statistics from report data"""
    stats = []
    
    if report_type == 'stock_summary':
        stats.extend([
            {
                'label': 'Total Products',
                'value': report_data.get('total_products', 0),
                'icon': 'package',
                'color': 'primary'
            },
            {
                'label': 'Total Stock Value',
                'value': f"{report_data.get('total_stock_value', 0):,.0f}",
                'icon': 'dollar-sign',
                'color': 'success'
            },
            {
                'label': 'Low Stock Items',
                'value': report_data.get('total_low_stock', 0),
                'icon': 'alert-triangle',
                'color': 'warning'
            },
        ])
    
    elif report_type == 'low_stock':
        stats.extend([
            {
                'label': 'Total Alerts',
                'value': report_data.get('total_alerts', 0),
                'icon': 'alert-circle',
                'color': 'danger'
            },
            {
                'label': 'Critical Alerts',
                'value': report_data.get('critical_alerts', 0),
                'icon': 'alert-triangle',
                'color': 'warning'
            },
        ])
    
    elif report_type == 'store_performance':
        if report_data.get('total_stores'):
            stats.append({
                'label': 'Stores',
                'value': report_data.get('total_stores', 0),
                'icon': 'store',
                'color': 'info'
            })
    
    return stats


def generate_insights(report_data, report_type):
    """Generate insights based on report data"""
    insights = []
    
    if report_type == 'low_stock' and report_data.get('critical_alerts', 0) > 0:
        insights.append({
            'title': 'Immediate Action Required',
            'message': f'{report_data["critical_alerts"]} products are critically low on stock.',
            'type': 'danger',
            'icon': 'alert-triangle',
            'action': 'View Details',
            'action_url': '#low-stock-section'
        })
    
    if report_type == 'stock_summary':
        low_stock_count = report_data.get('total_low_stock', 0)
        if low_stock_count > 0:
            insights.append({
                'title': 'Reordering Required',
                'message': f'{low_stock_count} products need reordering.',
                'type': 'warning',
                'icon': 'shopping-cart'
            })
    
    return insights


def generate_chart_data(data, report_type):
    """Generate chart data from report data"""
    if not data or not isinstance(data, list) or len(data) == 0:
        return None
    
    try:
        if report_type in ['stock_summary', 'inventory_valuation']:
            # Bar chart for top items by value
            sorted_data = sorted(data, key=lambda x: x.get('stock_value', 0), reverse=True)[:10]
            labels = [item.get('product_name', f'Item {i}')[:20] for i, item in enumerate(sorted_data)]
            values = [item.get('stock_value', 0) for item in sorted_data]
            
            return {
                'labels': labels,
                'datasets': [{
                    'label': 'Stock Value',
                    'data': values,
                    'backgroundColor': 'rgba(54, 162, 235, 0.5)',
                    'borderColor': 'rgba(54, 162, 235, 1)',
                    'borderWidth': 1
                }]
            }
        
        elif report_type == 'abc_analysis':
            # Pie chart for ABC classification
            abc_counts = {'A': 0, 'B': 0, 'C': 0}
            for item in data:
                abc_class = item.get('abc_class', 'C')
                if abc_class in abc_counts:
                    abc_counts[abc_class] += 1
            
            return {
                'labels': ['A Items (High Value)', 'B Items (Medium Value)', 'C Items (Low Value)'],
                'datasets': [{
                    'label': 'ABC Classification',
                    'data': [abc_counts['A'], abc_counts['B'], abc_counts['C']],
                    'backgroundColor': [
                        'rgba(255, 99, 132, 0.5)',
                        'rgba(54, 162, 235, 0.5)',
                        'rgba(255, 205, 86, 0.5)'
                    ],
                    'borderColor': [
                        'rgb(255, 99, 132)',
                        'rgb(54, 162, 235)',
                        'rgb(255, 205, 86)'
                    ],
                    'borderWidth': 1
                }]
            }
    
    except Exception as e:
        print(f"Error generating chart data: {e}")
        return None


@login_required
def export_report(request, report_type, format):
    """Export report to various formats"""
    
    # Generate report data
    report_methods = {
        'stock_summary': InventoryReports.generate_stock_summary_report,
        'low_stock': InventoryReports.generate_low_stock_alert_report,
        'store_performance': lambda: InventoryReports.generate_store_performance_report(None),
        'category_analysis': InventoryReports.generate_category_analysis_report,
        'product_movement': lambda: InventoryReports.generate_product_movement_report(None, 30),
        'abc_analysis': InventoryReports.generate_abc_analysis_report,
        'inventory_valuation': InventoryReports.generate_inventory_valuation_report,
        'stock_transfer': InventoryReports.generate_stock_transfer_report(None),
        'product_availability': InventoryReports.generate_product_availability_report(None),
    }
    
    if report_type not in report_methods:
        return HttpResponse('Invalid report type', status=400)
    
    try:
        report_data = report_methods[report_type]()
        
        if format == 'csv':
            response = HttpResponse(content_type='text/csv')
            filename = f"{report_type}_{timezone.now().strftime('%Y%m%d_%H%M%S')}.csv"
            response['Content-Disposition'] = f'attachment; filename="{filename}"'
            
            writer = csv.writer(response)
            
            # Write headers
            if report_data.get('data') and len(report_data['data']) > 0:
                headers = list(report_data['data'][0].keys())
                writer.writerow(headers)
                
                # Write data rows
                for row in report_data['data']:
                    writer.writerow([row.get(header, '') for header in headers])
            
            return response
        
        elif format == 'json':
            return JsonResponse(report_data, safe=False)
        
        else:
            return HttpResponse('Format not supported', status=400)
            
    except Exception as e:
        return HttpResponse(f'Error generating export: {str(e)}', status=500)